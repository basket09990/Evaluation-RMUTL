from django.shortcuts import render, redirect, get_object_or_404
from django.http.response import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from app_user.models import academic_type,wl_field,group_detail,Profile,group,user_evaluation_agreement,wl_subfield,user_competency_main,user_work_info,evr_round,main_competency,user_competency_main,specific_competency,user_competency_councilde,administrative_competency,user_competency_ceo,user_evaluation,user_evident
from django.http import HttpRequest, JsonResponse , HttpResponseRedirect
from django.contrib import messages 
from datetime import datetime,date
from django.db.models import Q
from django.core.paginator import Paginator
from django.urls import reverse_lazy,reverse
from django.views.generic import CreateView, UpdateView, DeleteView, ListView 
from app_user.models import WorkLeave,PersonalDiagram,UserWorkloadSelection,UserSelectedSubField,SelectedWorkload,SelectedSubfields,WorkloadCriteria,main_competency,Profile ,user_evaluation_score,UserMainCompetencyScore,UserAdministrativeCompetencyScore,UserSpecificCompetencyScore
from .forms import UserWorkloadSelectionForm,UserWorkInfoForm, GroupForm, GroupDetailForm, WlFieldForm ,WorkloadCriteriaForm,WlSubfieldForm ,MainCompetencyForm,SpecificCompetencyForm,AdministrativeCompetencyForm,GroupSelectionForm,UserEvaluationForm ,UserEvidentForm
from .forms import UserWorkloadSelectionForm1,UserEvaluationScoreForm , SubFieldForm , SelectSubfieldForm , WorkloadCriteriaSelectionForm, SubFieldSelectionForm,PersonalDiagramForm,WorkLeaveForm
from django.utils import timezone
from app_user.forms import UserProfileForm,ExtendedProfileForm
from django.forms import modelformset_factory
import uuid
from django.views.decorators.csrf import csrf_exempt
from PIL import Image
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import os
from io import BytesIO
from django.conf import settings
from django.db.models import Q
from django.db.models import Max
from django.db.models.signals import post_delete
from django.dispatch import receiver
import openpyxl
from django.template.loader import render_to_string
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Border, Side


# สัญญาณที่จะถูกเรียกเมื่อ UserSelectedSubField ถูกลบ
@receiver(post_delete, sender=UserSelectedSubField)
def delete_related_workload_selection(sender, instance, **kwargs):
    # ลบ UserWorkloadSelection ทั้งหมดที่มี sf_id ตรงกับที่ถูกลบ
    UserWorkloadSelection.objects.filter(
        user=instance.user,
        evaluation=instance.evaluation,
        sf_id=instance.sf_id
    ).delete()


# Views สำหรับจัดการ Group
def group_list(request):
    groups = group.objects.all()
    return render(request, 'app_evaluation/group_list.html', {'groups': groups})

def group_add(request):
    if request.method == 'POST':
        form = GroupForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('group_list')
    else:
        form = GroupForm()
    return render(request, 'app_evaluation/group_form.html', {'form': form, 'action': 'เพิ่มกลุ่ม'})

def group_edit(request, pk):
    group_instance = get_object_or_404(group, pk=pk)
    if request.method == 'POST':
        form = GroupForm(request.POST, instance=group_instance)
        if form.is_valid():
            form.save()
            return redirect('group_list')
    else:
        form = GroupForm(instance=group_instance)
    return render(request, 'app_evaluation/group_form.html', {'form': form, 'action': 'แก้ไขกลุ่ม'})

def group_delete(request, pk):
    group_instance = get_object_or_404(group, pk=pk)
    if request.method == 'POST':
        group_instance.delete()
        return redirect('group_list')
    return render(request, 'app_evaluation/group_confirm_delete.html', {'group': group_instance})

# Views สำหรับจัดการ GroupDetail
def group_detail_list(request, group_id):
    group_instance = get_object_or_404(group, pk=group_id)
    details = group_detail.objects.filter(g_id=group_instance)
    return render(request, 'app_evaluation/group_detail_list.html', {'group': group_instance, 'details': details})

def group_detail_add(request, group_id):
    group_instance = get_object_or_404(group, pk=group_id)

    if request.method == 'POST':
        form = GroupDetailForm(request.POST)
        if form.is_valid():
            detail = form.save(commit=False)
            detail.g_id = group_instance
            detail.save()
            return redirect('group_detail_list', group_id=group_instance.g_id)
    else:
        form = GroupDetailForm()
    
    return render(request, 'app_evaluation/group_detail_form.html', {
        'form': form, 
        'action': 'เพิ่มรายละเอียดกลุ่ม', 
        'group': group_instance
    })

def group_detail_edit(request, pk):
    detail = get_object_or_404(group_detail, pk=pk)
    if request.method == 'POST':
        form = GroupDetailForm(request.POST, instance=detail)
        if form.is_valid():
            form.save()
            return redirect('group_detail_list', group_id=detail.g_id.g_id)
    else:
        form = GroupDetailForm(instance=detail)
    return render(request, 'app_evaluation/group_detail_form.html', {'form': form, 'action': 'แก้ไขรายละเอียดกลุ่ม', 'group': detail.g_id})

def group_detail_delete(request, pk):
    detail = get_object_or_404(group_detail, pk=pk)
    if request.method == 'POST':
        group_id = detail.g_id.g_id
        detail.delete()
        return redirect('group_detail_list', group_id=group_id)
    return render(request, 'app_evaluation/group_detail_confirm_delete.html', {'detail': detail})

# Views สำหรับ wl_field
def wl_field_list(request):
    fields = wl_field.objects.all()
    return render(request, 'app_evaluation/wl_field_list.html', {'fields': fields})

def wl_field_add(request):
    if request.method == 'POST':
        form = WlFieldForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('wl_field_list')
    else:
        form = WlFieldForm()
    return render(request, 'app_evaluation/wl_field_form.html', {'form': form, 'action': 'เพิ่มกลุ่มภาระงาน'})

def wl_field_edit(request, pk):
    field = get_object_or_404(wl_field, pk=pk)
    if request.method == 'POST':
        form = WlFieldForm(request.POST, instance=field)
        if form.is_valid():
            form.save()
            return redirect('wl_field_list')
    else:
        form = WlFieldForm(instance=field)
    return render(request, 'app_evaluation/wl_field_form.html', {'form': form, 'action': 'แก้ไขกลุ่มภาระงาน'})

def wl_field_delete(request, pk):
    field = get_object_or_404(wl_field, pk=pk)
    if request.method == 'POST':
        field.delete()
        return redirect('wl_field_list')
    return render(request, 'app_evaluation/wl_field_confirm_delete.html', {'field': field})

# Views สำหรับ wl_subfield
def wl_subfield_list(request, pk):  # ใช้ 'pk' แทน 'field_id'
    field = get_object_or_404(wl_field, pk=pk)
    subfields = field.subfields.all()  # ใช้ related_name 'subfields' ที่กำหนดใน ForeignKey ของ wl_subfield
    return render(request, 'app_evaluation/wl_subfield_list.html', {
        'field': field,
        'subfields': subfields,
    })

def wl_subfield_add(request, field_id):
    field = get_object_or_404(wl_field, pk=field_id)
    if request.method == 'POST':
        form = WlSubfieldForm(request.POST)
        if form.is_valid():
            subfield = form.save(commit=False)
            subfield.f_id = field  # เชื่อม f_id กับ wl_field ที่ได้จาก field_id
            subfield.save()
            return redirect('wl_subfield_list', pk=field_id)  # ต้องส่ง pk ให้ครบ
    else:
        form = WlSubfieldForm()

    return render(request, 'app_evaluation/wl_subfield_form.html', {
        'form': form,
        'action': 'เพิ่มกลุ่มย่อย',
        'field': field,
        'is_edit': False
    })

def wl_subfield_edit(request, pk):
    subfield = get_object_or_404(wl_subfield, pk=pk)
    field = subfield.f_id  # เชื่อมกับ wl_field

    if request.method == 'POST':
        form = WlSubfieldForm(request.POST, instance=subfield)
        if form.is_valid():
            form.save()
            # ใช้ field.f_id.pk เพื่อส่งค่าที่ถูกต้อง
            return redirect('wl_subfield_list', pk=field.f_id)
    else:
        form = WlSubfieldForm(instance=subfield)

    return render(request, 'app_evaluation/wl_subfield_form.html', {
        'form': form,
        'action': 'แก้ไขกลุ่มย่อยภาระงาน',
        'field': field,
        'is_edit': True
    })

def wl_subfield_delete(request, pk):
    subfield = get_object_or_404(wl_subfield, pk=pk)
    field_id = subfield.f_id.pk  # ดึงค่า pk ของ wl_field ที่เกี่ยวข้อง

    if request.method == 'POST':
        subfield.delete()
        return redirect('wl_subfield_list', pk=field_id)  # ส่ง field_id กลับไป

    return render(request, 'app_evaluation/wl_subfield_confirm_delete.html', {
        'subfield': subfield
    })

# Views สำหรับ WorkloadCriteria
def workload_criteria_list(request, subfield_id):
    # ดึงข้อมูล wl_subfield ที่เกี่ยวข้อง
    subfield = get_object_or_404(wl_subfield, pk=subfield_id)

    # Query เพื่อดึง WorkloadCriteria ที่เชื่อมโยงกับ wl_subfield นี้
    criteria = WorkloadCriteria.objects.filter(sf_id=subfield)

    return render(request, 'app_evaluation/workload_criteria_list.html', {
        'subfield': subfield,
        'criteria': criteria,
    })

def workload_criteria_add(request, subfield_id):
    subfield = get_object_or_404(wl_subfield, pk=subfield_id)

    if request.method == 'POST':
        form = WorkloadCriteriaForm(request.POST)
        if form.is_valid():
            criteria = form.save(commit=False)
            criteria.sf_id = subfield  # เชื่อมกับ wl_subfield

            # ตรวจสอบว่ามีค่า f_id หรือไม่ ถ้าไม่มีกำหนดค่าให้ f_id
            if not criteria.f_id_id:  # ถ้ายังไม่มีการตั้งค่า f_id
                # สามารถตั้งค่า f_id ได้ เช่น:
                default_field = wl_field.objects.first()  # ดึงค่า wl_field แรกจากฐานข้อมูล
                if default_field:
                    criteria.f_id = default_field
                else:
                    # กรณีที่ไม่มี wl_field ในฐานข้อมูลเลย
                    messages.error(request, "ไม่พบข้อมูลฟิลด์ที่ถูกต้องในระบบ กรุณาตรวจสอบ")
                    return render(request, 'app_evaluation/workload_criteria_form.html', {
                        'form': form,
                        'subfield': subfield,
                        'action': 'เพิ่มเกณฑ์ภาระงาน',
                        'is_edit': False,
                        'subfield_id': subfield_id,  # ส่ง subfield_id ไปยัง template
                    })

            criteria.save()
            return redirect('workload_criteria_list', subfield_id=subfield.pk)  # ใช้ subfield.pk ที่ถูกต้อง
    else:
        form = WorkloadCriteriaForm()

    return render(request, 'app_evaluation/workload_criteria_form.html', {
        'form': form,
        'subfield': subfield,
        'action': 'เพิ่มเกณฑ์ภาระงาน',
        'is_edit': False,
        'subfield_id': subfield_id,  # ส่ง subfield_id ไปยัง template
    })

def workload_criteria_edit(request, pk):
    criteria = get_object_or_404(WorkloadCriteria, pk=pk)
    subfield_id = criteria.sf_id.sf_id  # ตรวจสอบให้แน่ใจว่าได้ subfield_id ที่ถูกต้อง

    if request.method == 'POST':
        form = WorkloadCriteriaForm(request.POST, instance=criteria)
        if form.is_valid():
            form.save()
            return redirect('workload_criteria_list', subfield_id=subfield_id)  # ส่ง subfield_id กลับไปใน redirect
    else:
        form = WorkloadCriteriaForm(instance=criteria)

    return render(request, 'app_evaluation/workload_criteria_form.html', {
        'form': form,
        'action': 'แก้ไขเกณฑ์ภาระงาน',
        'subfield_id': subfield_id,  # ส่ง subfield_id ไปยังเทมเพลต
        'is_edit': True,  # บอกว่าเป็นการแก้ไข
    })

def workload_criteria_delete(request, pk):
    criteria = get_object_or_404(WorkloadCriteria, pk=pk)

    if request.method == 'POST':
        subfield_id = criteria.sf_id.sf_id  # เก็บ subfield_id จาก criteria
        criteria.delete()
        # ส่ง subfield_id ใน redirect ให้ถูกต้อง
        return redirect('workload_criteria_list', subfield_id=subfield_id)

    return render(request, 'app_evaluation/workload_criteria_confirm_delete.html', {
        'criteria': criteria
    })

# แสดงรายการ main_competency
def list_main_competency(request):
    competencies = main_competency.objects.all()
    return render(request, 'app_evaluation/list_main_competency.html', {'competencies': competencies})

def add_main_competency(request):
    if request.method == 'POST':
        form = MainCompetencyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "เพิ่มข้อมูลสำเร็จ!")
            return redirect('list_main_competency')
    else:
        form = MainCompetencyForm()
    
    return render(request, 'app_evaluation/add_main_competency.html', {'form': form})

# แก้ไขข้อมูล
def edit_main_competency(request, mc_id):
    competency = get_object_or_404(main_competency, mc_id=mc_id)
    if request.method == 'POST':
        form = MainCompetencyForm(request.POST, instance=competency)
        if form.is_valid():
            form.save()
            messages.success(request, "แก้ไขข้อมูลสำเร็จ!")
            return redirect('list_main_competency')
    else:
        form = MainCompetencyForm(instance=competency)

    return render(request, 'app_evaluation/edit_main_competency.html', {'form': form, 'competency': competency})

# ลบข้อมูล
def delete_main_competency(request, mc_id):
    competency = get_object_or_404(main_competency, mc_id=mc_id)
    if request.method == 'POST':
        competency.delete()
        messages.success(request, "ลบข้อมูลสำเร็จ!")
        return redirect('list_main_competency')

    return render(request, 'app_evaluation/delete_main_competency.html', {'competency': competency})

# แสดงรายการ specific_competency
def list_specific_competency(request):
    competencies = specific_competency.objects.all()
    return render(request, 'app_evaluation/list_specific_competency.html', {'competencies': competencies})

# เพิ่มข้อมูล
def add_specific_competency(request):
    if request.method == 'POST':
        form = SpecificCompetencyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "เพิ่มข้อมูลสำเร็จ!")
            return redirect('list_specific_competency')
    else:
        form = SpecificCompetencyForm()
    
    return render(request, 'app_evaluation/add_specific_competency.html', {'form': form})

# แก้ไขข้อมูล
def edit_specific_competency(request, sc_id):
    competency = get_object_or_404(specific_competency, sc_id=sc_id)
    if request.method == 'POST':
        form = SpecificCompetencyForm(request.POST, instance=competency)
        if form.is_valid():
            form.save()
            messages.success(request, "แก้ไขข้อมูลสำเร็จ!")
            return redirect('list_specific_competency')
    else:
        form = SpecificCompetencyForm(instance=competency)

    return render(request, 'app_evaluation/edit_specific_competency.html', {'form': form, 'competency': competency})

# ลบข้อมูล
def delete_specific_competency(request, sc_id):
    competency = get_object_or_404(specific_competency, sc_id=sc_id)
    if request.method == 'POST':
        competency.delete()
        messages.success(request, "ลบข้อมูลสำเร็จ!")
        return redirect('list_specific_competency')

    return render(request, 'app_evaluation/delete_specific_competency.html', {'competency': competency})

# แสดงรายการ administrative_competency
def list_administrative_competency(request):
    competencies = administrative_competency.objects.all()
    return render(request, 'app_evaluation/list_administrative_competency.html', {'competencies': competencies})

# เพิ่มข้อมูล
def add_administrative_competency(request):
    if request.method == 'POST':
        form = AdministrativeCompetencyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "เพิ่มข้อมูลสำเร็จ!")
            return redirect('list_administrative_competency')
    else:
        form = AdministrativeCompetencyForm()
    
    return render(request, 'app_evaluation/add_administrative_competency.html', {'form': form})

# แก้ไขข้อมูล
def edit_administrative_competency(request, adc_id):
    competency = get_object_or_404(administrative_competency, adc_id=adc_id)
    if request.method == 'POST':
        form = AdministrativeCompetencyForm(request.POST, instance=competency)
        if form.is_valid():
            form.save()
            messages.success(request, "แก้ไขข้อมูลสำเร็จ!")
            return redirect('list_administrative_competency')
    else:
        form = AdministrativeCompetencyForm(instance=competency)

    return render(request, 'app_evaluation/edit_administrative_competency.html', {'form': form, 'competency': competency})

# ลบข้อมูล
def delete_administrative_competency(request, adc_id):
    competency = get_object_or_404(administrative_competency, adc_id=adc_id)
    if request.method == 'POST':
        competency.delete()
        messages.success(request, "ลบข้อมูลสำเร็จ!")
        return redirect('list_administrative_competency')

    return render(request, 'app_evaluation/delete_administrative_competency.html', {'competency': competency})

@login_required
def score(request):
    return render(request,'app_evaluation/score.html')

@login_required
def score0(request):
    return render(request,'app_evaluation/score0.html')

# ฟังก์ชันเพื่อดึงข้อมูลรอบการประเมิน
def get_evr_round():
    current_month = timezone.now().month
    current_year = timezone.now().year

    if 10 <= current_month <= 12:
        # รอบแรก (ปลายปี)
        round_number = 1
        evr_year = current_year
    elif 1 <= current_month <= 3:
        # รอบแรก (ต้นปีถัดไป)
        round_number = 1
        evr_year = current_year - 1
    else:
        # รอบสอง (กลางปี)
        round_number = 2
        evr_year = current_year

    evr_round_obj, created = evr_round.objects.get_or_create(
        evr_year=evr_year,
        evr_round=round_number,
    )
    return evr_round_obj

@login_required
def search_evaluation(request):
    # ดึงเฉพาะ evaluation ที่เชื่อมกับผู้ใช้ที่ล็อกอินอยู่ และเรียงลำดับจากใหม่ไปเก่า
    evaluations = user_evaluation.objects.filter(user=request.user).order_by('-uevr_id')
    
    query = request.GET.get('query', '')
    year = request.GET.get('year', '')
    round = request.GET.get('round', '')

    # กรองตามคำค้นหาชื่อผู้ใช้ ถ้ามีการป้อนข้อมูล
    if query:
        evaluations = evaluations.filter(user__username__icontains=query)
    
    # กรองตามปี ถ้ามีการป้อนข้อมูล
    if year:
        evaluations = evaluations.filter(evr_id__evr_year=year)
    
    # กรองตามรอบ ถ้ามีการป้อนข้อมูล
    if round:
        evaluations = evaluations.filter(evr_id__evr_round=round)

    # ดึงปีและรอบทั้งหมดเพื่อนำไปใช้ในตัวเลือก
    years = user_evaluation.objects.filter(user=request.user).values_list('evr_id__evr_year', flat=True).distinct()
    rounds = user_evaluation.objects.filter(user=request.user).values_list('evr_id__evr_round', flat=True).distinct()

    context = {
        'evaluations': evaluations,
        'query': query,
        'years': years,
        'rounds': rounds,
    }
    return render(request, 'app_evaluation/search_evaluations.html', context)

@login_required
def evaluation_page1(request, evaluation_id):
    user = request.user

    # Get the user evaluation object for the given evaluation_id
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id, user=user)
    evr_round_obj = evaluation.evr_id  # Get the current evaluation round from user_evaluation

    # Get the profile associated with the evaluation
    profile = evaluation.user.profile

    # Get the selected group from the user evaluation agreement
    user_evaluation_agreement_obj = user_evaluation_agreement.objects.filter(user=user, evr_id=evr_round_obj).first()
    selected_group = user_evaluation_agreement_obj.g_id if user_evaluation_agreement_obj else None

    # Get user work info
    user_work_info_obj = user_work_info.objects.filter(user=evaluation.user, round=evr_round_obj).first()
    work_form_current = UserWorkInfoForm(instance=user_work_info_obj)

    # Profile forms
    profile_form = UserProfileForm(instance=user)
    extended_form = ExtendedProfileForm(instance=profile)

    # Fetch the leave records for the current round and round 1
    def get_or_create_leave(user, round_obj, leave_type):
        leave, _ = WorkLeave.objects.get_or_create(
            user=user,
            round=round_obj,
            leave_type=leave_type,
            defaults={'times': 0, 'days': 0}
        )
        return leave

    # Get round 1 information (if any)
    round_1 = evr_round.objects.filter(
        evr_round=1, evr_year=(evr_round_obj.evr_year - 1 if evr_round_obj.evr_round == 2 else evr_round_obj.evr_year)
    ).first()

    # Create leave forms for round 1 and the current round
    sick_leave_round_1 = get_or_create_leave(user, round_1, 'SL') if round_1 else None
    sick_leave_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(user, evr_round_obj, 'SL'), prefix='sick_leave')
    
    personal_leave_round_1 = get_or_create_leave(user, round_1, 'PL') if round_1 else None
    personal_leave_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(user, evr_round_obj, 'PL'), prefix='personal_leave')

    late_round_1 = get_or_create_leave(user, round_1, 'LT') if round_1 else None
    late_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(user, evr_round_obj, 'LT'), prefix='late')

    maternity_leave_round_1 = get_or_create_leave(user, round_1, 'ML') if round_1 else None
    maternity_leave_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(user, evr_round_obj, 'ML'), prefix='maternity_leave')

    ordination_leave_round_1 = get_or_create_leave(user, round_1, 'OL') if round_1 else None
    ordination_leave_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(user, evr_round_obj, 'OL'), prefix='ordination_leave')

    longsick_leave_round_1 = get_or_create_leave(user, round_1, 'LSL') if round_1 else None
    longsick_leave_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(user, evr_round_obj, 'LSL'), prefix='longsick_leave')

    adsent_work_round_1 = get_or_create_leave(user, round_1, 'AW') if round_1 else None
    adsent_work_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(user, evr_round_obj, 'AW'), prefix='adsent_work')

    # Form lists for round 1 and the current round
    round_1_forms = {
        'sick_leave_round_1_form': WorkLeaveForm(request.POST or None, instance=sick_leave_round_1, prefix='sick_leave_round_1') if round_1 else None,
        'personal_leave_round_1_form': WorkLeaveForm(request.POST or None, instance=personal_leave_round_1, prefix='personal_leave_round_1') if round_1 else None,
        'late_round_1_form': WorkLeaveForm(request.POST or None, instance=late_round_1, prefix='late_round_1') if round_1 else None,
        'maternity_leave_round_1_form': WorkLeaveForm(request.POST or None, instance=maternity_leave_round_1, prefix='maternity_leave_round_1') if round_1 else None,
        'ordination_leave_round_1_form': WorkLeaveForm(request.POST or None, instance=ordination_leave_round_1, prefix='ordination_leave_round_1') if round_1 else None,
        'longsick_leave_round_1_form': WorkLeaveForm(request.POST or None, instance=longsick_leave_round_1, prefix='longsick_leave_round_1') if round_1 else None,
        'adsent_work_round_1_form': WorkLeaveForm(request.POST or None, instance=adsent_work_round_1, prefix='adsent_work_round_1') if round_1 else None
    }

    # Handle form submission
    if request.method == 'POST':
        # Validate all forms
        profile_form = UserProfileForm(request.POST, instance=user)
        extended_form = ExtendedProfileForm(request.POST, instance=profile)
        work_form_current = UserWorkInfoForm(request.POST, instance=user_work_info_obj)

        # Validate both round 1 forms and current round forms
        current_round_forms = [sick_leave_form, personal_leave_form, late_form, maternity_leave_form, ordination_leave_form, longsick_leave_form, adsent_work_form]
        all_forms = [profile_form, extended_form, work_form_current] + current_round_forms + list(round_1_forms.values())

        if all(form.is_valid() for form in all_forms if form is not None):
            # Save all forms
            profile_form.save()
            extended_form.save()
            work_form_current.save()
            
            for form in current_round_forms:
                form.save()

            for form in round_1_forms.values():
                if form:
                    form.save()

            messages.success(request, "บันทึกข้อมูลเรียบร้อยแล้ว!")
            return redirect('evaluation_page1', evaluation_id=evaluation_id)
        else:
            # Debugging: Print form errors
            for form in all_forms:
                print(form.errors)  # Check the errors for each form
            messages.error(request, "มีข้อผิดพลาดในการบันทึกข้อมูล กรุณาตรวจสอบข้อมูลอีกครั้ง")
    # Prepare the context for rendering
    context = {
        'profile_form': profile_form,
        'extended_form': extended_form,
        'work_form_current': work_form_current,
        'evr_round': evr_round_obj,
        'profile': profile,
        'selected_group': selected_group,
        'sick_leave_form': sick_leave_form,
        'personal_leave_form': personal_leave_form,
        'late_form': late_form,
        'maternity_leave_form': maternity_leave_form,
        'ordination_leave_form': ordination_leave_form,
        'longsick_leave_form': longsick_leave_form,
        'adsent_work_form': adsent_work_form,
        # Include round 1 forms in the context if they exist
        **round_1_forms,
        'evaluation_id': evaluation_id,
        'user_evaluation_agreement_year_thai': user_evaluation_agreement_obj.year + 543,
        'user_evaluation_agreement': user_evaluation_agreement_obj,
    }

    return render(request, 'app_evaluation/evaluation_page1.html', context)

@login_required
def evaluation_page2(request, evaluation_id):
    user = request.user
    evr_round_obj = get_evr_round()

    # ดึงข้อมูล user_evaluation ที่สัมพันธ์กับ evaluation_id
    try:
        user_evaluation_obj = user_evaluation.objects.get(pk=evaluation_id, user=user)
    except user_evaluation.DoesNotExist:
        messages.error(request, "ไม่พบข้อมูลการประเมิน")
        return redirect('select_group')

    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)


    # ดึง selected_group ที่สัมพันธ์กับ user ผ่าน user_evaluation_agreement
    try:
        user_evaluation_agreement_obj = user_evaluation_agreement.objects.get(user=user, evr_id=user_evaluation_obj.evr_id)
        selected_group = user_evaluation_agreement_obj.g_id
    except user_evaluation_agreement.DoesNotExist:
        selected_group = None

    # ดึงข้อมูล f_id ที่เชื่อมโยงกับ group_detail ของ selected_group
    if selected_group:
        fields = wl_field.objects.filter(group_detail__g_id=selected_group).distinct()
    else:
        fields = wl_field.objects.none()

    # ดึงข้อมูล subfield ที่ผู้ใช้เลือก
    selected_subfields = UserSelectedSubField.objects.filter( evaluation=evaluation)

    # ดึงข้อมูล evaluation
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)

    # ดึงข้อมูล workload_selections
    workload_selections = UserWorkloadSelection.objects.filter(evaluation=evaluation)

    # ดึงข้อมูล subfields ที่เกี่ยวข้องกับ fields ที่ได้เลือก
    subfields = wl_subfield.objects.filter(f_id__in=fields)

    # คำนวณค่า min_workload จากข้อมูล group
    min_workload = group.objects.filter(g_id=selected_group.g_id).values_list('g_max_workload', flat=True).first()
    if not min_workload:
        min_workload = 35  # ตั้งค่า default ถ้าไม่มีข้อมูล

    # คำนวณผลรวมของ calculated_workload
    total_workload = sum(selection.calculated_workload for selection in workload_selections) if workload_selections else 0

    # ดึง c_wl ที่สูงที่สุดจาก user_evaluation
    max_workload = user_evaluation.objects.filter( evr_id=user_evaluation_obj.evr_id).aggregate(max_workload=Max('c_wl'))['max_workload'] or 0

    # คำนวณส่วนต่างคะแนนภาระงาน
    workload_difference = max_workload - total_workload
    workload_difference_score = workload_difference * (28 / 115)
    achievement_work = 70 - workload_difference_score

    # บันทึกค่าที่คำนวณใน evaluation
    evaluation.c_wl = total_workload
    evaluation.achievement_work = round(achievement_work, 2)
    evaluation.save()

    # สร้าง formset สำหรับ WorkloadCriteria
    WorkloadFormset = modelformset_factory(WorkloadCriteria, form=WorkloadCriteriaForm, extra=0)

    # สร้าง formsets สำหรับแต่ละ field
    formsets = {}
    for field in fields:
        formset = WorkloadFormset(queryset=WorkloadCriteria.objects.filter(f_id=field), prefix=f'field_{field.f_id}')
        formsets[field.f_id] = formset

    # เมื่อผู้ใช้ทำการ POST ข้อมูล
    if request.method == 'POST':
        is_valid = True
        # วนลูปตรวจสอบและบันทึกข้อมูลในแต่ละ formset
        for field_id, formset in formsets.items():
            filled_formset = WorkloadFormset(request.POST, queryset=WorkloadCriteria.objects.filter(f_id__f_id=field_id), prefix=f'field_{field_id}')
            if filled_formset.is_valid():
                filled_formset.save()
            else:
                is_valid = False  # ถ้ามีฟอร์มไหนไม่ถูกต้อง ให้ตั้งค่านี้เป็น False

        if is_valid:
            messages.success(request, 'บันทึกข้อมูลเรียบร้อยแล้ว!')
            return redirect('evaluation_page2', evaluation_id=evaluation_id)
        else:
            messages.error(request, 'มีข้อผิดพลาดในการบันทึกข้อมูล กรุณาตรวจสอบข้อมูลอีกครั้ง')

    # ส่ง context ไปยังเทมเพลต
    context = {
        'user_evaluation': user_evaluation_obj,
        'formsets': formsets,
        'fields': fields,
        'selected_subfields': selected_subfields,
        'evaluation': evaluation,
        'workload_selections': workload_selections,
        'subfields': subfields,
        'total_workload': total_workload,
        'achievement_work': evaluation.achievement_work,
        'evaluation_id': evaluation_id,
        'min_workload': min_workload,
        'user_evaluation_agreement_year_thai': user_evaluation_agreement_obj.year + 543,  # ค่า year_thai ที่ส่งมาจาก page1
    }

    return render(request, 'app_evaluation/evaluation_page2.html', context)

@login_required
def select_subfields2(request, f_id, evaluation_id):
    field = get_object_or_404(wl_field, pk=f_id)
    subfields = wl_subfield.objects.filter(f_id=field)

    # ดึงข้อมูลการประเมินที่สอดคล้องกับ evaluation_id
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)

    if request.method == 'POST':
        selected_sf_id = request.POST.get('sf_id')  # ใช้ get() เพื่อดึงค่า sf_id จากฟอร์ม
        if selected_sf_id:
            selected_sf = wl_subfield.objects.get(sf_id=selected_sf_id)
            # เพิ่มข้อมูล evaluation เข้าไปในการสร้าง UserSelectedSubField
            UserSelectedSubField.objects.create(
                user=request.user,
                evaluation=evaluation,  # เชื่อม evaluation
                f_id=field,
                sf_id=selected_sf
            )
            return redirect('evaluation_page2', evaluation_id=evaluation_id)

    context = {
        'field': field,
        'subfields': subfields,
        'evaluation_id': evaluation_id,
    }
    return render(request, 'app_evaluation/select_subfields2.html', context)

@login_required
def delete_selected_subfield2(request, sf_id):
    # ตรวจสอบว่ามีการส่ง evaluation_id มาหรือไม่
    evaluation_id = request.POST.get('evaluation_id') or request.GET.get('evaluation_id')
    if not evaluation_id:
        messages.error(request, "ไม่พบ evaluation_id")
        return redirect('select_group')

    try:
        # ลบข้อมูลทั้งหมดใน UserSelectedSubField ที่มี sf_id
        selected_subfields = UserSelectedSubField.objects.filter(sf_id=sf_id)

        if not selected_subfields.exists():
            messages.error(request, 'ไม่พบ Subfield ที่คุณพยายามจะลบ.')
            return redirect('evaluation_page2', evaluation_id=evaluation_id)

        if request.method == 'POST':
            # ลบข้อมูลที่ค้นหาได้ทั้งหมด
            selected_subfields.delete()
            messages.success(request, 'ลบ Subfield เรียบร้อยแล้ว!')
            return redirect('evaluation_page2', evaluation_id=evaluation_id)

    except Exception as e:
        messages.error(request, f'เกิดข้อผิดพลาด: {str(e)}')
        return redirect('evaluation_page2', evaluation_id=evaluation_id)

    return redirect('evaluation_page2', evaluation_id=evaluation_id)

@login_required
def select_workload_criteria2(request, evaluation_id, sf_id):
    # ดึงข้อมูล evaluation และ subfield ที่เกี่ยวข้อง
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    subfield = get_object_or_404(wl_subfield, pk=sf_id)
    
    # สร้างฟอร์มและส่งค่า subfield ไปยังฟอร์มเพื่อกรองข้อมูลใน selected_id
    criteria_form = UserWorkloadSelectionForm(initial={'sf_id': subfield}, subfield=subfield)

    # ตรวจสอบการส่งฟอร์ม
    if request.method == 'POST':
        criteria_form = UserWorkloadSelectionForm(request.POST, subfield=subfield)
        if criteria_form.is_valid():
            new_criteria = criteria_form.save(commit=False)
            new_criteria.sf_id = subfield
            new_criteria.evaluation = evaluation
            new_criteria.user = request.user

            try:
                # แปลงค่าทั้งหมดเป็น float ก่อนทำการคำนวณ
                selected_num = float(new_criteria.selected_num )
                selected_maxnum = float(new_criteria.selected_maxnum )
                selected_workload = float(new_criteria.selected_workload )

                # คำนวณค่า calculated_workload
                if selected_maxnum == 0:
                    calculated_workload = selected_num * selected_workload
                else:
                    if selected_num <= selected_maxnum:
                        calculated_workload = selected_num * selected_workload
                    else:
                        calculated_workload = selected_maxnum* selected_workload

                # แสดงข้อมูลการคำนวณใน log
                print(f"selected_num: {selected_num}, selected_maxnum: {selected_maxnum}, selected_workload: {selected_workload}")
                print(f"calculated_workload: {calculated_workload}")

                # บันทึกค่า calculated_workload ลงใน new_criteria
                new_criteria.calculated_workload = calculated_workload
                new_criteria.save()

                messages.success(request, f'เพิ่มภาระงานใหม่เรียบร้อยแล้ว! คำนวณภาระงาน: {calculated_workload:.2f}')
                return redirect('evaluation_page2', evaluation_id=evaluation_id)

            except ValueError as e:
                # ถ้ามีข้อผิดพลาดในการแปลงค่า แสดงใน log
                print(f"ValueError: {e}")
                messages.error(request, 'ไม่สามารถคำนวณค่าได้ โปรดตรวจสอบข้อมูลที่กรอก.')

    # ส่งข้อมูลไปยังเทมเพลต
    context = {
        'evaluation': evaluation,
        'subfield': subfield,
        'criteria_form': criteria_form,
    }
    return render(request, 'app_evaluation/select_workload_criteria2.html', context)

@login_required
def edit_workload_selection2(request, selection_id):
    selection = get_object_or_404(UserWorkloadSelection, pk=selection_id)

    if request.method == 'POST':
        form = UserWorkloadSelectionForm1(request.POST, instance=selection)

        if form.is_valid():
            # ถ้าฟอร์มถูกต้อง ให้บันทึกการเปลี่ยนแปลง
            if selection.selected_maxnum == 0:
                selection.calculated_workload = selection.selected_num * selection.selected_workload
            else:
                if selection.selected_num <= selection.selected_maxnum:
                    selection.calculated_workload = selection.selected_num * selection.selected_workload
                else:
                    selection.calculated_workload = selection.selected_maxnum  * selection.selected_workload
            form.save()
            messages.success(request, 'แก้ไขภาระงานเรียบร้อยแล้ว!')
            return redirect('evaluation_page2', evaluation_id=selection.evaluation.uevr_id)
        else:
            messages.error(request, 'กรุณาเลือกข้อมูลภาระงานให้ครบถ้วน.')
    else:
        form = UserWorkloadSelectionForm1(instance=selection)

    return render(request, 'app_evaluation/edit_workload_selection2.html', {
        'form': form,
        'selection': selection,
    })

@login_required
def delete_workload_selection2(request, selection_id):
    selection = get_object_or_404(UserWorkloadSelection, pk=selection_id)

    if request.method == 'POST':
        # ลบเฉพาะการเชื่อมโยงกับ UserWorkloadSelection
        selection.delete()  
        messages.success(request, 'ลบภาระงานเรียบร้อยแล้ว!')
        return redirect('evaluation_page2', evaluation_id=selection.evaluation.uevr_id)

    return render(request, 'app_evaluation/confirm_delete2.html', {
        'selection': selection,
    })

@login_required
def upload_evidence1(request, criteria_id):
    # ดึงข้อมูล WorkloadCriteria โดยใช้ criteria_id
    criteria = get_object_or_404(UserWorkloadSelection, pk=criteria_id)
    evaluation = criteria.evaluation  # ดึงค่า evaluation จาก UserWorkloadSelection

    if request.method == 'POST':
        form = UserEvidentForm(request.POST, request.FILES)
        files = request.FILES.getlist('file')
        pictures = request.FILES.getlist('picture')

        if form.is_valid():
            # บันทึกไฟล์ PDF และ DOCX
            for file in files:
                new_filename = f"{uuid.uuid4()}_{file.name}"
                evidence = user_evident(
                    uwls_id=criteria,  # ตั้งค่า uwls_id ด้วย instance ของ UserWorkloadSelection
                    uevr_id=evaluation,  # ตั้งค่า uevr_id ด้วย instance ของ UserEvaluation
                    file=file,
                    filename=new_filename
                )
                evidence.save()

            # ลดขนาดรูปภาพก่อนบันทึก
            for picture in pictures:
                new_filename = f"{uuid.uuid4()}_{picture.name}"
                image = Image.open(picture)

                # ลดขนาดรูปภาพ
                max_size = (800, 800)  # ขนาดที่ต้องการ
                image.thumbnail(max_size)

                # บันทึกรูปภาพที่ลดขนาด
                img_io = ContentFile(b'')
                image_format = image.format or 'JPEG'
                image.save(img_io, format=image_format)

                # บันทึกไฟล์ลงใน storage
                file_name = default_storage.save(f"uploads/{new_filename}", img_io)
                evidence = user_evident(
                    uwls_id=criteria,  # ตั้งค่า uwls_id ด้วย instance ของ UserWorkloadSelection
                    uevr_id=evaluation,  # ตั้งค่า uevr_id ด้วย instance ของ UserEvaluation
                    picture=file_name,
                    filename=new_filename
                )
                evidence.save()

            messages.success(request, "อัปโหลดไฟล์เรียบร้อยแล้ว!")
            return redirect('upload_evidence1', criteria_id=criteria_id)

    else:
        form = UserEvidentForm()

    # แสดงรายการไฟล์ที่อัปโหลด
    evidences = user_evident.objects.filter(uwls_id=criteria)

    return render(request, 'app_evaluation/upload_evidence1.html', {
        'form': form,
        'evaluation': evaluation,
        'evidences': evidences  # ส่งรายการไฟล์/รูปภาพไปยัง template
    })


@login_required
def delete_evidence1(request, evidence_id):
    evidence = get_object_or_404(user_evident, uevd_id=evidence_id)

    # ลบไฟล์หรือรูปภาพที่เกี่ยวข้องจากระบบ
    if evidence.file:
        if os.path.isfile(evidence.file.path):
            os.remove(evidence.file.path)  # ลบไฟล์จาก storage
    if evidence.picture:
        if os.path.isfile(evidence.picture.path):
            os.remove(evidence.picture.path)  # ลบรูปภาพจาก storage

    # ลบรายการ evidence จากฐานข้อมูล
    evidence.delete()
    messages.success(request, "ลบไฟล์/รูปภาพเรียบร้อยแล้ว!")
    
    # ใช้ `redirect` ให้ไปยัง `upload_evidence` โดยใช้ `criteria_id`
    return redirect('upload_evidence1', criteria_id=evidence.uwls_id.id)  # เปลี่ยน `evaluation_id` เป็น `criteria_id`


def calculate_competency_score(actual_score, expected_score):
    # ตรวจสอบว่า actual_score และ expected_score ไม่ใช่ None
    if actual_score is None or expected_score is None:
        return 0

    # คำนวณความแตกต่างระหว่าง actual_score (คะแนนที่กรอก) และ expected_score (ระดับที่คาดหวัง)
    difference = expected_score - actual_score

    if difference <= 0:  # สมรรถนะที่แสดงออก >= คาดหวัง
        return 3  # คูณด้วย 3 คะแนน
    elif difference == 1:  # ต่ำกว่าคาดหวัง 1 ระดับ
        return 2  # คูณด้วย 2 คะแนน
    elif difference == 2:  # ต่ำกว่าคาดหวัง 2 ระดับ
        return 1  # คูณด้วย 1 คะแนน
    else:  # ต่ำกว่าคาดหวัง 3 ระดับหรือมากกว่า
        return 0  # คูณด้วย 0 คะแนน


@login_required
def evaluation_page3(request, evaluation_id):
    # ดึงข้อมูล user_evaluation
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    

    # ดึงข้อมูล profile ที่เชื่อมกับ user_evaluation
    profile = evaluation.user.profile

    # ดึงข้อมูลการประเมินผลของผู้ใช้
    user_evaluation_agreement_obj = user_evaluation_agreement.objects.filter(user=evaluation.user).first()
    selected_group = user_evaluation_agreement_obj.g_id if user_evaluation_agreement_obj else None

    main_competencies = main_competency.objects.filter(mc_type=evaluation.ac_id.ac_name)
    specific_competencies = specific_competency.objects.filter(sc_type=evaluation.ac_id.ac_name)
    administrative_competencies = None  # ตั้งค่าเริ่มต้นเป็น None
    
    # เช็คว่าถ้ามีตำแหน่งบริหารและไม่ใช่ "-"
    if evaluation.administrative_position and evaluation.administrative_position != "-":
        administrative_competencies = administrative_competency.objects.filter()

    # ตรวจสอบการส่งข้อมูล POST
    if request.method == 'POST':
        # วนลูปผ่าน main_competencies
        for competency in main_competencies:
            actual_score = request.POST.get(f'main_actual_score_{competency.mc_id}')
            if actual_score and actual_score.isdigit():
                actual_score = int(actual_score)
                user_competency_main.objects.update_or_create(
                    evaluation=evaluation,
                    mc_id=competency,
                  
                    defaults={'user': evaluation.user, 'umc_name': competency.mc_name, 'umc_type': competency.mc_type, 'actual_score': actual_score}
                )

        # วนลูปผ่าน specific_competencies
        for competency in specific_competencies:
            actual_score = request.POST.get(f'specific_actual_score_{competency.sc_id}')
            if actual_score and actual_score.isdigit():
                actual_score = int(actual_score)
                user_competency_councilde.objects.update_or_create(
                    evaluation=evaluation,
                    sc_id=competency,
                   
                    defaults={'user': evaluation.user, 'ucc_name': competency.sc_name, 'ucc_type': competency.sc_type, 'actual_score': actual_score}
                )

        # วนลูปผ่าน administrative_competencies
        if administrative_competencies is not None:
            for competency in administrative_competencies:
                actual_score = request.POST.get(f'admin_actual_score_{competency.adc_id}')
                uceo_num = request.POST.get(f'admin_uceo_num_{competency.adc_id}') 
                if actual_score and actual_score.isdigit():
                    actual_score = int(actual_score)
                    uceo_num = int(uceo_num) if uceo_num.isdigit() else 0
                    user_competency_ceo.objects.update_or_create(
                        evaluation=evaluation,
                        adc_id=competency,
                        
                        defaults={'user': evaluation.user, 'uceo_name': competency.adc_name, 'uceo_type': competency.adc_type, 'actual_score': actual_score,'uceo_num': uceo_num}
                    )

        messages.success(request, "บันทึกคะแนนเรียบร้อยแล้ว!")
        return redirect('evaluation_page3', evaluation_id=evaluation_id)

    # ดึงข้อมูลคะแนนที่เคยกรอก
    main_scores = user_competency_main.objects.filter(evaluation=evaluation)
    specific_scores = user_competency_councilde.objects.filter(evaluation=evaluation)
    administrative_scores = user_competency_ceo.objects.filter(evaluation=evaluation)

    # เก็บจำนวนสมรรถนะที่ได้คะแนนตามเงื่อนไขต่าง ๆ
    score_count = {3: 0, 
                   2: 0, 
                   1: 0, 
                   0: 0}

    # คำนวณคะแนนสำหรับ main competencies
    main_competency_total = 0
    for score in main_scores:
        calculated_score = calculate_competency_score(score.actual_score, score.mc_id.mc_num)  # ส่ง expected_score ด้วย
        score_count[calculated_score] += 1
        main_competency_total += calculated_score

    # คำนวณคะแนนสำหรับ specific competencies
    specific_competency_total = 0
    for score in specific_scores:
        calculated_score = calculate_competency_score(score.actual_score, score.sc_id.sc_num)  # ส่ง expected_score ด้วย
        score_count[calculated_score] += 1
        specific_competency_total += calculated_score

    # คำนวณคะแนนสำหรับ administrative competencies (ถ้ามี)
    administrative_competency_total = 0
    if administrative_competencies is not None:
        for score in administrative_scores:
            calculated_score = calculate_competency_score(score.actual_score, score.uceo_num)
            score_count[calculated_score] += 1
            administrative_competency_total += calculated_score

    # ผลรวมคะแนนทั้งหมด
    total_score = main_competency_total + specific_competency_total + administrative_competency_total

    # คำนวณคะแนนในแต่ละกรณี (คูณ 3, 2, 1 และ 0)
    score_3_total = score_count[3] * 3
    score_2_total = score_count[2] * 2
    score_1_total = score_count[1] * 1
    score_0_total = score_count[0] * 0

    # คำนวณค่ารวมของคะแนนที่คาดหวังสำหรับแต่ละตาราง
    main_max_num = sum([c.mc_num for c in main_competencies])
    specific_max_num = sum([c.sc_num for c in specific_competencies])
    administrative_max_num = sum([score.uceo_num for score in administrative_scores]) if administrative_competencies is not None else 0

    # คำนวณคะแนนจากสูตร
    total_max_num = main_max_num + specific_max_num + administrative_max_num

    if total_max_num > 0:
        calculated_score = (total_score / total_max_num) * 30
    print(f"Total Score: {total_score}")
    print(f"Total Max Num: {total_max_num}")
    # บันทึก calculated_score ใน mc_score ของ user_evaluation
    evaluation.mc_score = calculated_score  # บันทึกคะแนนที่คำนวณได้
    evaluation.save()  # บันทึกการเปลี่ยนแปลงลงในฐานข้อมูล


    # ส่งข้อมูลไปยังเทมเพลต
    context = {
        'profile': profile,
        'main_competencies': main_competencies,
        'specific_competencies': specific_competencies,
        'administrative_competencies': administrative_competencies,
        'main_scores': main_scores,
        'specific_scores': specific_scores,
        'administrative_scores': administrative_scores,
        'main_competency_total': main_competency_total,
        'specific_competency_total': specific_competency_total,
        'administrative_competency_total': administrative_competency_total,
        'total_score': total_score,
        'main_max_num': main_max_num,
        'specific_max_num': specific_max_num,
        'administrative_max_num': administrative_max_num,
        'calculated_score': calculated_score,
        'evaluation_id': evaluation_id,
        'user_evaluation_obj': evaluation,
        'ac_name': evaluation.ac_id.ac_name,
        'count_3': score_count[3],
        'count_2': score_count[2],
        'count_1': score_count[1],
        'count_0': score_count[0],
        'score_3_total': score_count[3] * 3,
        'score_2_total': score_count[2] * 2,
        'score_1_total': score_count[1] * 1,
        'score_0_total': score_count[0] * 0,
        'user_evaluation_obj': evaluation,
    }

    return render(request, 'app_evaluation/evaluation_page3.html', context)

@login_required
def evaluation_page4(request, evaluation_id):
    # Get the current user
    user = request.user

    # Fetch the evaluation object based on ID and user
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id, user=user)
    evr_round_obj = evaluation.evr_id  # Get the current evaluation round from user_evaluation

    # ดึงข้อมูลโปรไฟล์ที่เชื่อมโยงกับการประเมินนี้
    profile = evaluation.user.profile

    # ตรวจสอบว่ามีการบันทึก `user_evaluation_agreement` ไว้หรือไม่
    user_evaluation_agreement_obj = user_evaluation_agreement.objects.filter(user=user, evr_id=evr_round_obj).first()
    selected_group = user_evaluation_agreement_obj.g_id if user_evaluation_agreement_obj else None

    # Fetch or create remarks fields
    if request.method == 'POST':
        # ดึงข้อมูลจากฟอร์มที่ถูกส่งมา
        remark_achievement = request.POST.get('remark_achievement', evaluation.remark_achievement)
        remark_mc = request.POST.get('remark_mc', evaluation.remark_mc)
        remark_other = request.POST.get('remark_other', evaluation.remark_other)
        remark_total = request.POST.get('remark_total', evaluation.remark_total)

        # บันทึกข้อมูลหมายเหตุ
        evaluation.remark_achievement = remark_achievement
        evaluation.remark_mc = remark_mc
        evaluation.remark_other = remark_other
        evaluation.remark_total = remark_total
        evaluation.save()  # บันทึกการเปลี่ยนแปลง

        messages.success(request, 'บันทึกหมายเหตุเรียบร้อยแล้ว!')
        return redirect('evaluation_page4', evaluation_id=evaluation_id)

    # Example: Calculate totals, aggregate scores, etc.
    achievement_work = evaluation.achievement_work or 0
    mc_score = evaluation.mc_score or 0
    total_score = achievement_work + mc_score

    print(f"Achievement Work: {achievement_work}, MC Score: {mc_score}, Total Score: {total_score}")

    if total_score >= 90:
        level = 'ดีเด่น'
    elif total_score >= 80:
        level = 'ดีมาก'
    elif total_score >= 70:
        level = 'ดี'
    elif total_score >= 60:
        level = 'พอใช้'
    else:
        level = 'ต้องปรับปรุง'

    print(f"Total Score: {total_score}, Level: {level}")

    # Prepare the context for rendering
    context = {
        'user_evaluation': evaluation,
        'achievement_work': achievement_work,  # Achievement work value
        'mc_score': mc_score,
        'total_score': total_score,
        'level': level,
        'remark_achievement': evaluation.remark_achievement,
        'remark_mc': evaluation.remark_mc,
        'remark_other': evaluation.remark_other,
        'remark_total': evaluation.remark_total,
        'evaluation_id': evaluation_id,
        'profile': profile,
        'selected_group': selected_group,
        # Add more variables as needed
    }

    return render(request, 'app_evaluation/evaluation_page4.html', context)

@login_required
def evaluation_page5(request, evaluation_id):
    # Fetch the evaluation object
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id, user=request.user)

    # Initialize form and formset
    PersonalDiagramFormset = modelformset_factory(PersonalDiagram, fields=('skill_evol', 'dev', 'dev_time'), extra=0, can_delete=True)

    
    form = UserEvaluationForm(instance=evaluation)

    # ตรวจสอบ method POST
    if request.method == 'POST':
        # โหลด formset จากข้อมูล POST ที่ส่งเข้ามา
        formset = PersonalDiagramFormset(request.POST)
        if 'save_form' in request.POST:
            form = UserEvaluationForm(request.POST, instance=evaluation)
            if form.is_valid():
                form.save()
                messages.success(request, "บันทึกข้อมูลการประเมินสำเร็จ!")
            else:
                messages.error(request, "เกิดข้อผิดพลาดในการบันทึกข้อมูล กรุณาตรวจสอบข้อมูลที่กรอก")
        
        elif 'save_formset' in request.POST:
            if formset.is_valid():
                formset.save()
                messages.success(request, "บันทึกข้อมูลฟอร์มสำเร็จ!")
            else:
                messages.error(request, "เกิดข้อผิดพลาดในการบันทึกข้อมูล กรุณาตรวจสอบข้อมูลที่กรอก")
        
        # โหลด formset ใหม่เพื่อรีเฟรชข้อมูล
        formset = PersonalDiagramFormset(queryset=PersonalDiagram.objects.filter(uevr_id=evaluation))

    else:
        # โหลด formset เมื่อ request เป็น GET
        formset = PersonalDiagramFormset(queryset=PersonalDiagram.objects.filter(uevr_id=evaluation))

    context = {
        'form': form,
        'formset': formset,
        'evaluation_id': evaluation_id,
    }
    return render(request, 'app_evaluation/evaluation_page5.html', context)

@login_required
def add_personal_diagram1(request, evaluation_id):
    # ดึง user_evaluation ตาม evaluation_id
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)

    if request.method == 'POST':
        form = PersonalDiagramForm(request.POST)
        if form.is_valid():
            personal_diagram = form.save(commit=False)
            personal_diagram.uevr_id = evaluation  # เชื่อมโยงกับ user_evaluation
            personal_diagram.save()
            messages.success(request, 'เพิ่มข้อมูลแผนพัฒนาสำเร็จ!')
            # กลับไปยังหน้า evaluation_page_5
            return redirect('evaluation_page5', evaluation_id=evaluation_id)
        else:
            messages.error(request, 'เกิดข้อผิดพลาดในการบันทึก กรุณาตรวจสอบข้อมูล.')
    else:
        form = PersonalDiagramForm()

    context = {
        'form': form,
        'evaluation_id': evaluation_id,
    }
    return render(request, 'app_evaluation/add_personal_diagram1.html', context)


@login_required
def edit_personal_diagram1(request, pd_id):
    # ดึงข้อมูล PersonalDiagram ที่ต้องการแก้ไข
    personal_diagram = get_object_or_404(PersonalDiagram, pk=pd_id)

    if request.method == 'POST':
        form = PersonalDiagramForm(request.POST, instance=personal_diagram)
        if form.is_valid():
            form.save()
            messages.success(request, 'แก้ไขข้อมูลแผนพัฒนาสำเร็จ!')
            # กลับไปยังหน้า evaluation_page_5
            return redirect('evaluation_page5', evaluation_id=personal_diagram.uevr_id.uevr_id)
        else:
            messages.error(request, 'เกิดข้อผิดพลาดในการแก้ไข กรุณาตรวจสอบข้อมูล.')
    else:
        form = PersonalDiagramForm(instance=personal_diagram)

    context = {
        'form': form,
        'evaluation_id': personal_diagram.uevr_id.uevr_id,
    }
    return render(request, 'app_evaluation/edit_personal_diagram1.html', context)

@login_required
def delete_personal_diagram1(request, pd_id):
    # ดึงข้อมูล PersonalDiagram ที่ต้องการลบ
    personal_diagram = get_object_or_404(PersonalDiagram, pk=pd_id)

    if request.method == 'POST':
        evaluation_id = personal_diagram.uevr_id.uevr_id
        
        try:
            # ตรวจสอบว่าค่า pd_id ถูกต้อง
            print(f"Trying to delete PersonalDiagram with ID: {pd_id}, Linked Evaluation ID: {evaluation_id}")
            personal_diagram.delete()
            messages.success(request, 'ลบข้อมูลแผนพัฒนาสำเร็จ!')
        except Exception as e:
            # แสดงข้อความข้อผิดพลาดที่เกิดขึ้น
            print(f"Error occurred while deleting PersonalDiagram ID {pd_id}: {e}")
            messages.error(request, f'เกิดข้อผิดพลาดในการลบข้อมูล: {e}')
        return redirect('evaluation_page5', evaluation_id=evaluation_id)
    
    return redirect('evaluation_page5', evaluation_id=personal_diagram.uevr_id.uevr_id)

@login_required
def search_evaluations_2(request):
    query = request.GET.get('q')  # รับค่าค้นหาจาก URL เช่น ?q=คำค้นหา
    year = request.GET.get('year')  # รับค่าปีจาก URL
    round = request.GET.get('round')  # รับค่ารอบจาก URL
    
    evaluations = user_evaluation.objects.all()

    # ตรวจสอบว่ามีการค้นหาหรือไม่
    if query:
        evaluations = evaluations.filter(
            user__first_name__icontains=query
        ) | evaluations.filter(
            user__last_name__icontains=query
        )
    
    # ตรวจสอบว่ามีการกรองตามปีหรือไม่
    if year:
        evaluations = evaluations.filter(evr_id__evr_year=year)
    
    # ตรวจสอบว่ามีการกรองตามรอบหรือไม่
    if round:
        evaluations = evaluations.filter(evr_id__evr_round=round)

    # เรียงข้อมูลให้ใหม่สุดอยู่ด้านบน โดยใช้ฟิลด์ created_at ในการจัดเรียง
    evaluations = evaluations.order_by('-created_at')  # เปลี่ยน 'created_at' เป็นฟิลด์เวลาที่เหมาะสม

    # ดึงข้อมูลปีและรอบทั้งหมดเพื่อนำไปแสดงในฟอร์มค้นหา
    years = user_evaluation.objects.values_list('evr_id__evr_year', flat=True).distinct()
    rounds = user_evaluation.objects.values_list('evr_id__evr_round', flat=True).distinct()

    context = {
        'evaluations': evaluations,
        'query': query,
        'years': years,
        'rounds': rounds,
    }

    return render(request, 'app_evaluation/search_evaluations_2.html', context)

@login_required
def evaluation_page_from_1(request, evaluation_id):
    # ดึงข้อมูล user_evaluation โดยไม่สนใจ request.user
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    evr_round_obj = evaluation.evr_id  # Get the current evaluation round from user_evaluation

    # ดึงข้อมูล profile ที่เชื่อมกับ user_evaluation
    profile = evaluation.user.profile

    # ดึงข้อมูลการประเมินผลของผู้ใช้
    user_evaluation_agreement_obj = user_evaluation_agreement.objects.filter(user=evaluation.user, evr_id=evr_round_obj).first()
    selected_group = user_evaluation_agreement_obj.g_id if user_evaluation_agreement_obj else None

    # ดึงข้อมูลการปฏิบัติงานของผู้ใช้
    user_work_info_obj, created = user_work_info.objects.get_or_create(
        user=evaluation.user,
        round=evr_round_obj,
        defaults={'punishment': ''}
    )
    work_form_current = UserWorkInfoForm(instance=user_work_info_obj)

    # สร้างฟอร์มโปรไฟล์และการปฏิบัติงาน
    profile_form = UserProfileForm(instance=evaluation.user)
    extended_form = ExtendedProfileForm(instance=profile)

    # ฟังก์ชันสำหรับดึงข้อมูลการลา
    def get_or_create_leave(user, round_obj, leave_type):
        leave, _ = WorkLeave.objects.get_or_create(
            user=user,
            round=round_obj,
            leave_type=leave_type,
            defaults={'times': 0, 'days': 0}
        )
        return leave

    # ตรวจสอบรอบที่ 1 (หากมีข้อมูล)
    round_1 = evr_round.objects.filter(
        evr_round=1, evr_year=(evr_round_obj.evr_year - 1 if evr_round_obj.evr_round == 2 else evr_round_obj.evr_year)
    ).first()

    # ดึงข้อมูลการลาของรอบที่ 1 และรอบปัจจุบัน
    sick_leave_round_1 = get_or_create_leave(evaluation.user, round_1, 'SL') if round_1 else None
    sick_leave_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(evaluation.user, evr_round_obj, 'SL'), prefix='sick_leave')
    
    personal_leave_round_1 = get_or_create_leave(evaluation.user, round_1, 'PL') if round_1 else None
    personal_leave_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(evaluation.user, evr_round_obj, 'PL'), prefix='personal_leave')

    late_round_1 = get_or_create_leave(evaluation.user, round_1, 'LT') if round_1 else None
    late_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(evaluation.user, evr_round_obj, 'LT'), prefix='late')

    maternity_leave_round_1 = get_or_create_leave(evaluation.user, round_1, 'ML') if round_1 else None
    maternity_leave_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(evaluation.user, evr_round_obj, 'ML'), prefix='maternity_leave')

    ordination_leave_round_1 = get_or_create_leave(evaluation.user, round_1, 'OL') if round_1 else None
    ordination_leave_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(evaluation.user, evr_round_obj, 'OL'), prefix='ordination_leave')

    longsick_leave_round_1 = get_or_create_leave(evaluation.user, round_1, 'LSL') if round_1 else None
    longsick_leave_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(evaluation.user, evr_round_obj, 'LSL'), prefix='longsick_leave')

    adsent_work_round_1 = get_or_create_leave(evaluation.user, round_1, 'AW') if round_1 else None
    adsent_work_form = WorkLeaveForm(request.POST or None, instance=get_or_create_leave(evaluation.user, evr_round_obj, 'AW'), prefix='adsent_work')

    # ฟอร์มสำหรับรอบที่ 1 และรอบปัจจุบัน
    round_1_forms = {
        'sick_leave_round_1_form': WorkLeaveForm(request.POST or None, instance=sick_leave_round_1, prefix='sick_leave_round_1') if round_1 else None,
        'personal_leave_round_1_form': WorkLeaveForm(request.POST or None, instance=personal_leave_round_1, prefix='personal_leave_round_1') if round_1 else None,
        'late_round_1_form': WorkLeaveForm(request.POST or None, instance=late_round_1, prefix='late_round_1') if round_1 else None,
        'maternity_leave_round_1_form': WorkLeaveForm(request.POST or None, instance=maternity_leave_round_1, prefix='maternity_leave_round_1') if round_1 else None,
        'ordination_leave_round_1_form': WorkLeaveForm(request.POST or None, instance=ordination_leave_round_1, prefix='ordination_leave_round_1') if round_1 else None,
        'longsick_leave_round_1_form': WorkLeaveForm(request.POST or None, instance=longsick_leave_round_1, prefix='longsick_leave_round_1') if round_1 else None,
        'adsent_work_round_1_form': WorkLeaveForm(request.POST or None, instance=adsent_work_round_1, prefix='adsent_work_round_1') if round_1 else None
    }

    # จัดการการบันทึกข้อมูลฟอร์ม
    if request.method == 'POST':
        # ตรวจสอบข้อมูลทุกฟอร์ม
        profile_form = UserProfileForm(request.POST, instance=evaluation.user)
        extended_form = ExtendedProfileForm(request.POST, instance=profile)
        work_form_current = UserWorkInfoForm(request.POST, instance=user_work_info_obj)

        # ฟอร์มทั้งหมดที่จะตรวจสอบความถูกต้อง
        current_round_forms = [sick_leave_form, personal_leave_form, late_form, maternity_leave_form, ordination_leave_form, longsick_leave_form, adsent_work_form]
        all_forms = [profile_form, extended_form, work_form_current] + current_round_forms + list(round_1_forms.values())

        if all(form.is_valid() for form in all_forms if form is not None):
            # บันทึกข้อมูลทุกฟอร์ม
            profile_form.save()
            extended_form.save()
            work_form_current.save()
            
            for form in current_round_forms:
                form.save()

            for form in round_1_forms.values():
                if form:
                    form.save()

            messages.success(request, "บันทึกข้อมูลเรียบร้อยแล้ว!")
            return redirect('evaluation_page_from_1', evaluation_id=evaluation_id)
        else:
            messages.error(request, "มีข้อผิดพลาดในการบันทึกข้อมูล กรุณาตรวจสอบข้อมูลอีกครั้ง")

    # เตรียมข้อมูลเพื่อส่งไปแสดงผล
    context = {
        'profile_form': profile_form,
        'extended_form': extended_form,
        'work_form_current': work_form_current,
        'evr_round': evr_round_obj,
        'profile': profile,
        'selected_group': selected_group,
        'sick_leave_form': sick_leave_form,
        'personal_leave_form': personal_leave_form,
        'late_form': late_form,
        'maternity_leave_form': maternity_leave_form,
        'ordination_leave_form': ordination_leave_form,
        'longsick_leave_form': longsick_leave_form,
        'adsent_work_form': adsent_work_form,
        # ฟอร์มรอบที่ 1 ถ้ามี
        **round_1_forms,
        'evaluation_id': evaluation_id,
        'user_evaluation_agreement_year_thai': user_evaluation_agreement_obj.year + 543 if user_evaluation_agreement_obj else None,
        'user_evaluation_agreement': user_evaluation_agreement_obj,
        'user_evaluation': evaluation,
    }

    return render(request, 'app_evaluation/evaluation_page_from_1.html', context)


@login_required
def evaluation_page_from_2(request, evaluation_id):
    # ดึงข้อมูล user_evaluation โดยไม่สนใจ request.user
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    evr_round_obj = evaluation.evr_id

    # ดึงข้อมูล profile
    profile = evaluation.user.profile

    # ดึง selected_group ที่สัมพันธ์กับ user ผ่าน user_evaluation_agreement
    try:
        user_evaluation_agreement_obj = user_evaluation_agreement.objects.get(user=evaluation.user, evr_id=evr_round_obj)
        selected_group = user_evaluation_agreement_obj.g_id
    except user_evaluation_agreement.DoesNotExist:
        selected_group = None

    # ดึงข้อมูล f_id ที่เชื่อมโยงกับ group_detail ของ selected_group
    if selected_group:
        fields = wl_field.objects.filter(group_detail__g_id=selected_group).distinct()
    else:
        fields = wl_field.objects.none()

    # ดึงข้อมูล subfields ที่ผู้ใช้เลือก
    selected_subfields = UserSelectedSubField.objects.filter(evaluation=evaluation)

    # ดึงข้อมูล workload_selections
    workload_selections = UserWorkloadSelection.objects.filter(evaluation=evaluation)

    # ดึงข้อมูล subfields ที่เกี่ยวข้องกับ fields ที่ได้เลือก
    subfields = wl_subfield.objects.filter(f_id__in=fields)

    # คำนวณค่า min_workload จากข้อมูล group
    min_workload = group.objects.filter(g_id=selected_group.g_id).values_list('g_max_workload', flat=True).first()
    if not min_workload:
        min_workload = 35  # ตั้งค่า default ถ้าไม่มีข้อมูล

    # คำนวณผลรวมของ calculated_workload
    total_workload = sum(selection.calculated_workload for selection in workload_selections) if workload_selections else 0

    # ดึง c_wl ที่สูงที่สุดจาก user_evaluation
    max_workload = user_evaluation.objects.filter(evr_id=evaluation.evr_id).aggregate(max_workload=Max('c_wl'))['max_workload'] or 0

    # คำนวณส่วนต่างคะแนนภาระงาน
    workload_difference = max_workload - total_workload
    workload_difference_score = workload_difference * (28 / 115)
    achievement_work = 70 - workload_difference_score

    # บันทึกค่าที่คำนวณใน evaluation
    evaluation.c_wl = total_workload
    evaluation.achievement_work = round(achievement_work, 2)
    evaluation.save()


    # ส่ง context ไปยังเทมเพลต
    context = {
        'user_evaluation': evaluation,
        'fields': fields,
        'selected_subfields': selected_subfields,
        'workload_selections': workload_selections,
        'subfields': subfields,
        'total_workload': total_workload,
        'achievement_work': evaluation.achievement_work,
        'evaluation_id': evaluation_id,
        'min_workload': min_workload,
        'user_evaluation_agreement_year_thai': user_evaluation_agreement_obj.year + 543 if user_evaluation_agreement_obj else None,

    }

    return render(request, 'app_evaluation/evaluation_page_from_2.html', context)


@login_required
def select_subfields3(request, f_id, evaluation_id):
    field = get_object_or_404(wl_field, pk=f_id)
    subfields = wl_subfield.objects.filter(f_id=field)

    # ดึงข้อมูลการประเมินที่สอดคล้องกับ evaluation_id
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)

    if request.method == 'POST':
        selected_sf_id = request.POST.get('sf_id')  # ใช้ get() เพื่อดึงค่า sf_id จากฟอร์ม
        if selected_sf_id:
            selected_sf = wl_subfield.objects.get(sf_id=selected_sf_id)
            # เพิ่มข้อมูล evaluation เข้าไปในการสร้าง UserSelectedSubField
            UserSelectedSubField.objects.create(
                user=request.user,
                evaluation=evaluation,  # เชื่อม evaluation
                f_id=field,
                sf_id=selected_sf
            )
            return redirect('evaluation_page_from_2', evaluation_id=evaluation_id)

    context = {
        'field': field,
        'subfields': subfields,
        'evaluation_id': evaluation_id,
    }
    return render(request, 'app_evaluation/select_subfields3.html', context)

@login_required
def delete_selected_subfield3(request, sf_id):
    # ตรวจสอบว่ามีการส่ง evaluation_id มาหรือไม่
    evaluation_id = request.POST.get('evaluation_id') or request.GET.get('evaluation_id')
    if not evaluation_id:
        messages.error(request, "ไม่พบ evaluation_id")
        return redirect('select_group')

    try:
        # ลบข้อมูลทั้งหมดใน UserSelectedSubField ที่มี sf_id
        selected_subfields = UserSelectedSubField.objects.filter(sf_id=sf_id)

        if not selected_subfields.exists():
            messages.error(request, 'ไม่พบ Subfield ที่คุณพยายามจะลบ.')
            return redirect('evaluation_page_from_2', evaluation_id=evaluation_id)

        if request.method == 'POST':
            # ลบข้อมูลที่ค้นหาได้ทั้งหมด
            selected_subfields.delete()
            messages.success(request, 'ลบ Subfield เรียบร้อยแล้ว!')
            return redirect('evaluation_page_from_2', evaluation_id=evaluation_id)

    except Exception as e:
        messages.error(request, f'เกิดข้อผิดพลาด: {str(e)}')
        return redirect('evaluation_page_from_2', evaluation_id=evaluation_id)

    return redirect('evaluation_page_from_2', evaluation_id=evaluation_id)


@login_required
def select_workload_criteria3(request, evaluation_id, sf_id):
    # ดึงข้อมูล evaluation และ subfield ที่เกี่ยวข้อง
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    subfield = get_object_or_404(wl_subfield, pk=sf_id)
    
    # สร้างฟอร์มและส่งค่า subfield ไปยังฟอร์มเพื่อกรองข้อมูลใน selected_id
    criteria_form = UserWorkloadSelectionForm(initial={'sf_id': subfield}, subfield=subfield)

    # ตรวจสอบการส่งฟอร์ม
    if request.method == 'POST':
        criteria_form = UserWorkloadSelectionForm(request.POST, subfield=subfield)
        if criteria_form.is_valid():
            new_criteria = criteria_form.save(commit=False)
            new_criteria.sf_id = subfield
            new_criteria.evaluation = evaluation
            new_criteria.user = request.user

            try:
                # แปลงค่าทั้งหมดเป็น float ก่อนทำการคำนวณ
                selected_num = float(new_criteria.selected_num )
                selected_maxnum = float(new_criteria.selected_maxnum )
                selected_workload = float(new_criteria.selected_workload )

                # คำนวณค่า calculated_workload
                if selected_maxnum == 0:
                    calculated_workload = selected_num  * selected_workload
                else:
                    if selected_num <= selected_maxnum:
                        calculated_workload = selected_num  * selected_workload
                    else:
                        calculated_workload = selected_maxnum  * selected_workload

                # แสดงข้อมูลการคำนวณใน log
                print(f"selected_num: {selected_num}, selected_maxnum: {selected_maxnum}, selected_workload: {selected_workload}")
                print(f"calculated_workload: {calculated_workload}")

                # บันทึกค่า calculated_workload ลงใน new_criteria
                new_criteria.calculated_workload = calculated_workload
                new_criteria.save()

                messages.success(request, f'เพิ่มภาระงานใหม่เรียบร้อยแล้ว! คำนวณภาระงาน: {calculated_workload:.2f}')
                return redirect('evaluation_page_from_2', evaluation_id=evaluation_id)

            except ValueError as e:
                # ถ้ามีข้อผิดพลาดในการแปลงค่า แสดงใน log
                print(f"ValueError: {e}")
                messages.error(request, 'ไม่สามารถคำนวณค่าได้ โปรดตรวจสอบข้อมูลที่กรอก.')

    # ส่งข้อมูลไปยังเทมเพลต
    context = {
        'evaluation': evaluation,
        'subfield': subfield,
        'criteria_form': criteria_form,
    }
    return render(request, 'app_evaluation/select_workload_criteria3.html', context)

@login_required
def edit_workload_selection3(request, selection_id):
    selection = get_object_or_404(UserWorkloadSelection, pk=selection_id)

    if request.method == 'POST':
        form = UserWorkloadSelectionForm1(request.POST, instance=selection)

        if form.is_valid():
            # ถ้าฟอร์มถูกต้อง ให้บันทึกการเปลี่ยนแปลง
            if selection.selected_maxnum == 0:
                selection.calculated_workload = selection.selected_num * selection.selected_workload
            else:
                if selection.selected_num <= selection.selected_maxnum:
                    selection.calculated_workload = selection.selected_num  * selection.selected_workload
                else:
                    selection.calculated_workload = selection.selected_maxnum * selection.selected_workload
            form.save()
            messages.success(request, 'แก้ไขภาระงานเรียบร้อยแล้ว!')
            return redirect('evaluation_page_from_2', evaluation_id=selection.evaluation.uevr_id)
        else:
            messages.error(request, 'กรุณาเลือกข้อมูลภาระงานให้ครบถ้วน.')
    else:
        form = UserWorkloadSelectionForm1(instance=selection)

    return render(request, 'app_evaluation/edit_workload_selection3.html', {
        'form': form,
        'selection': selection,
    })

@login_required
def delete_workload_selection3(request, selection_id):
    selection = get_object_or_404(UserWorkloadSelection, pk=selection_id)

    if request.method == 'POST':
        # ลบเฉพาะการเชื่อมโยงกับ UserWorkloadSelection
        selection.delete()  
        messages.success(request, 'ลบภาระงานเรียบร้อยแล้ว!')
        return redirect('evaluation_page_from_2', evaluation_id=selection.evaluation.uevr_id)

    return render(request, 'app_evaluation/confirm_delete3.html', {
        'selection': selection,
    })

@login_required
def upload_evidence2(request, criteria_id):
    # ดึงข้อมูล WorkloadCriteria โดยใช้ criteria_id
    criteria = get_object_or_404(UserWorkloadSelection, pk=criteria_id)
    evaluation = criteria.evaluation  # ดึงค่า evaluation จาก UserWorkloadSelection

    if request.method == 'POST':
        form = UserEvidentForm(request.POST, request.FILES)
        files = request.FILES.getlist('file')
        pictures = request.FILES.getlist('picture')

        if form.is_valid():
            # บันทึกไฟล์ PDF และ DOCX
            for file in files:
                new_filename = f"{uuid.uuid4()}_{file.name}"
                evidence = user_evident(
                    uwls_id=criteria,  # ตั้งค่า uwls_id ด้วย instance ของ UserWorkloadSelection
                    uevr_id=evaluation,  # ตั้งค่า uevr_id ด้วย instance ของ UserEvaluation
                    file=file,
                    filename=new_filename
                )
                evidence.save()

            # ลดขนาดรูปภาพก่อนบันทึก
            for picture in pictures:
                new_filename = f"{uuid.uuid4()}_{picture.name}"
                image = Image.open(picture)

                # ลดขนาดรูปภาพ
                max_size = (800, 800)  # ขนาดที่ต้องการ
                image.thumbnail(max_size)

                # บันทึกรูปภาพที่ลดขนาด
                img_io = ContentFile(b'')
                image_format = image.format or 'JPEG'
                image.save(img_io, format=image_format)

                # บันทึกไฟล์ลงใน storage
                file_name = default_storage.save(f"uploads/{new_filename}", img_io)
                evidence = user_evident(
                    uwls_id=criteria,  # ตั้งค่า uwls_id ด้วย instance ของ UserWorkloadSelection
                    uevr_id=evaluation,  # ตั้งค่า uevr_id ด้วย instance ของ UserEvaluation
                    picture=file_name,
                    filename=new_filename
                )
                evidence.save()

            messages.success(request, "อัปโหลดไฟล์เรียบร้อยแล้ว!")
            return redirect('upload_evidence2', criteria_id=criteria_id)

    else:
        form = UserEvidentForm()

    # แสดงรายการไฟล์ที่อัปโหลด
    evidences = user_evident.objects.filter(uwls_id=criteria)

    return render(request, 'app_evaluation/upload_evidence2.html', {
        'form': form,
        'evaluation': evaluation,
        'evidences': evidences  # ส่งรายการไฟล์/รูปภาพไปยัง template
    })


@login_required
def delete_evidence2(request, evidence_id):
    evidence = get_object_or_404(user_evident, uevd_id=evidence_id)

    # ลบไฟล์หรือรูปภาพที่เกี่ยวข้องจากระบบ
    if evidence.file:
        if os.path.isfile(evidence.file.path):
            os.remove(evidence.file.path)  # ลบไฟล์จาก storage
    if evidence.picture:
        if os.path.isfile(evidence.picture.path):
            os.remove(evidence.picture.path)  # ลบรูปภาพจาก storage

    # ลบรายการ evidence จากฐานข้อมูล
    evidence.delete()
    messages.success(request, "ลบไฟล์/รูปภาพเรียบร้อยแล้ว!")
    
    # ใช้ `redirect` ให้ไปยัง `upload_evidence` โดยใช้ `criteria_id`
    return redirect('upload_evidence2', criteria_id=evidence.uwls_id.id)  # เปลี่ยน `evaluation_id` เป็น `criteria_id`

@login_required
def evaluation_page_from_3(request, evaluation_id):
    # ดึงข้อมูล user_evaluation
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    

    # ดึงข้อมูล profile ที่เชื่อมกับ user_evaluation
    profile = evaluation.user.profile

    # ดึงข้อมูลการประเมินผลของผู้ใช้
    user_evaluation_agreement_obj = user_evaluation_agreement.objects.filter(user=evaluation.user).first()
    selected_group = user_evaluation_agreement_obj.g_id if user_evaluation_agreement_obj else None

    main_competencies = main_competency.objects.filter(mc_type=evaluation.ac_id.ac_name)
    specific_competencies = specific_competency.objects.filter(sc_type=evaluation.ac_id.ac_name)
    administrative_competencies = None  # ตั้งค่าเริ่มต้นเป็น None
    
    # เช็คว่าถ้ามีตำแหน่งบริหารและไม่ใช่ "-"
    if evaluation.administrative_position and evaluation.administrative_position != "-":
        administrative_competencies = administrative_competency.objects.filter()

    # ตรวจสอบการส่งข้อมูล POST
    if request.method == 'POST':
        # วนลูปผ่าน main_competencies
        for competency in main_competencies:
            actual_score = request.POST.get(f'main_actual_score_{competency.mc_id}')
            if actual_score and actual_score.isdigit():
                actual_score = int(actual_score)
                user_competency_main.objects.update_or_create(
                    evaluation=evaluation,
                    mc_id=competency,
                  
                    defaults={'user': evaluation.user, 'umc_name': competency.mc_name, 'umc_type': competency.mc_type, 'actual_score': actual_score}
                )

        # วนลูปผ่าน specific_competencies
        for competency in specific_competencies:
            actual_score = request.POST.get(f'specific_actual_score_{competency.sc_id}')
            if actual_score and actual_score.isdigit():
                actual_score = int(actual_score)
                user_competency_councilde.objects.update_or_create(
                    evaluation=evaluation,
                    sc_id=competency,
                   
                    defaults={'user': evaluation.user, 'ucc_name': competency.sc_name, 'ucc_type': competency.sc_type, 'actual_score': actual_score}
                )

        # วนลูปผ่าน administrative_competencies
        if administrative_competencies is not None:
            for competency in administrative_competencies:
                actual_score = request.POST.get(f'admin_actual_score_{competency.adc_id}')
                uceo_num = request.POST.get(f'admin_uceo_num_{competency.adc_id}') 
                if actual_score and actual_score.isdigit():
                    actual_score = int(actual_score)
                    uceo_num = int(uceo_num) if uceo_num.isdigit() else 0
                    user_competency_ceo.objects.update_or_create(
                        evaluation=evaluation,
                        adc_id=competency,
                        
                        defaults={'user': evaluation.user, 'uceo_name': competency.adc_name, 'uceo_type': competency.adc_type, 'actual_score': actual_score,'uceo_num': uceo_num}
                    )

        messages.success(request, "บันทึกคะแนนเรียบร้อยแล้ว!")
        return redirect('evaluation_page_from_3', evaluation_id=evaluation_id)

    # ดึงข้อมูลคะแนนที่เคยกรอก
    main_scores = user_competency_main.objects.filter(evaluation=evaluation)
    specific_scores = user_competency_councilde.objects.filter(evaluation=evaluation)
    administrative_scores = user_competency_ceo.objects.filter(evaluation=evaluation)

    # เก็บจำนวนสมรรถนะที่ได้คะแนนตามเงื่อนไขต่าง ๆ
    score_count = {3: 0, 
                   2: 0, 
                   1: 0, 
                   0: 0}

    # คำนวณคะแนนสำหรับ main competencies
    main_competency_total = 0
    for score in main_scores:
        calculated_score = calculate_competency_score(score.actual_score, score.mc_id.mc_num)  # ส่ง expected_score ด้วย
        score_count[calculated_score] += 1
        main_competency_total += calculated_score

    # คำนวณคะแนนสำหรับ specific competencies
    specific_competency_total = 0
    for score in specific_scores:
        calculated_score = calculate_competency_score(score.actual_score, score.sc_id.sc_num)  # ส่ง expected_score ด้วย
        score_count[calculated_score] += 1
        specific_competency_total += calculated_score

    # คำนวณคะแนนสำหรับ administrative competencies (ถ้ามี)
    administrative_competency_total = 0
    if administrative_competencies is not None:
        for score in administrative_scores:
            calculated_score = calculate_competency_score(score.actual_score, score.uceo_num)
            score_count[calculated_score] += 1
            administrative_competency_total += calculated_score

    # ผลรวมคะแนนทั้งหมด
    total_score = main_competency_total + specific_competency_total + administrative_competency_total

    # คำนวณคะแนนในแต่ละกรณี (คูณ 3, 2, 1 และ 0)
    score_3_total = score_count[3] * 3
    score_2_total = score_count[2] * 2
    score_1_total = score_count[1] * 1
    score_0_total = score_count[0] * 0

    # คำนวณค่ารวมของคะแนนที่คาดหวังสำหรับแต่ละตาราง
    main_max_num = sum([c.mc_num for c in main_competencies])
    specific_max_num = sum([c.sc_num for c in specific_competencies])
    administrative_max_num = sum([score.uceo_num for score in administrative_scores]) if administrative_competencies is not None else 0

    # คำนวณคะแนนจากสูตร
    total_max_num = main_max_num + specific_max_num + administrative_max_num

    if total_max_num > 0:
        calculated_score = (total_score / total_max_num) * 30
    print(f"Total Score: {total_score}")
    print(f"Total Max Num: {total_max_num}")
    # บันทึก calculated_score ใน mc_score ของ user_evaluation
    evaluation.mc_score = calculated_score  # บันทึกคะแนนที่คำนวณได้
    evaluation.save()  # บันทึกการเปลี่ยนแปลงลงในฐานข้อมูล


    # ส่งข้อมูลไปยังเทมเพลต
    context = {
        'profile': profile,
        'main_competencies': main_competencies,
        'specific_competencies': specific_competencies,
        'administrative_competencies': administrative_competencies,
        'main_scores': main_scores,
        'specific_scores': specific_scores,
        'administrative_scores': administrative_scores,
        'main_competency_total': main_competency_total,
        'specific_competency_total': specific_competency_total,
        'administrative_competency_total': administrative_competency_total,
        'total_score': total_score,
        'main_max_num': main_max_num,
        'specific_max_num': specific_max_num,
        'administrative_max_num': administrative_max_num,
        'calculated_score': calculated_score,
        'evaluation_id': evaluation_id,
        'user_evaluation_obj': evaluation,
        'ac_name': evaluation.ac_id.ac_name,
        'count_3': score_count[3],
        'count_2': score_count[2],
        'count_1': score_count[1],
        'count_0': score_count[0],
        'score_3_total': score_count[3] * 3,
        'score_2_total': score_count[2] * 2,
        'score_1_total': score_count[1] * 1,
        'score_0_total': score_count[0] * 0,
        'user_evaluation_obj': evaluation,
    }

    return render(request, 'app_evaluation/evaluation_page_from_3.html', context)


@login_required
def evaluation_page_from_4(request, evaluation_id):
    # ดึงข้อมูล user_evaluation
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)


    # ดึงข้อมูล profile ที่เชื่อมกับ user_evaluation
    profile = evaluation.user.profile

    # ดึงข้อมูลการประเมินผลของผู้ใช้
    user_evaluation_agreement_obj = user_evaluation_agreement.objects.filter(user=evaluation.user).first()
    selected_group = user_evaluation_agreement_obj.g_id if user_evaluation_agreement_obj else None

    # ตรวจสอบการร้องขอจากผู้ใช้
    if request.method == 'POST':
        # ดึงข้อมูลจากฟอร์มที่ถูกส่งมา
        remark_achievement = request.POST.get('remark_achievement')
        remark_mc = request.POST.get('remark_mc')
        remark_other = request.POST.get('remark_other')
        remark_total = request.POST.get('remark_total')

        # บันทึกข้อมูลหมายเหตุ
        evaluation.remark_achievement = remark_achievement
        evaluation.remark_mc = remark_mc
        evaluation.remark_other = remark_other
        evaluation.remark_total = remark_total

        evaluation.save()  # บันทึกการเปลี่ยนแปลง

        messages.success(request, 'บันทึกหมายเหตุเรียบร้อยแล้ว!')
        return redirect('evaluation_page_from_4', evaluation_id=evaluation_id)
    
    # คำนวณคะแนนต่าง ๆ
    achievement_work = evaluation.achievement_work or 0
    mc_score = evaluation.mc_score or 0
    total_score = achievement_work + mc_score

    # กำหนดระดับผลการประเมิน
    if total_score >= 90:
        level = 'ดีเด่น'
    elif total_score >= 80:
        level = 'ดีมาก'
    elif total_score >= 70:
        level = 'ดี'
    elif total_score >= 60:
        level = 'พอใช้'
    else:
        level = 'ต้องปรับปรุง'

    # ส่งข้อมูลไปยังเทมเพลต
    context = {
        'user_evaluation': evaluation,
        'achievement_work': achievement_work,  # Achievement work value
        'mc_score': mc_score,
        'total_score': total_score,
        'level': level,
        'remark_achievement': evaluation.remark_achievement,
        'remark_mc': evaluation.remark_mc,
        'remark_other': evaluation.remark_other,
        'remark_total': evaluation.remark_total,
        'evaluation_id': evaluation_id,
    }

    return render(request, 'app_evaluation/evaluation_page_from_4.html', context)

@login_required
def evaluation_page_from_5(request, evaluation_id):
    # ดึงข้อมูล user_evaluation โดยไม่สนใจ request.user
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    
    # ดึงข้อมูล profile ที่เชื่อมกับ user_evaluation
    profile = evaluation.user.profile

    # ดึงข้อมูลการประเมินผลของผู้ใช้
    user_evaluation_agreement_obj = user_evaluation_agreement.objects.filter(user=evaluation.user).first()
    selected_group = user_evaluation_agreement_obj.g_id if user_evaluation_agreement_obj else None

    # สร้างฟอร์ม PersonalDiagramFormset และ UserEvaluationForm
    PersonalDiagramFormset = modelformset_factory(PersonalDiagram, fields=('skill_evol', 'dev', 'dev_time'), extra=0)
    form = UserEvaluationForm(instance=evaluation)
    formset = PersonalDiagramFormset(queryset=PersonalDiagram.objects.filter(uevr_id=evaluation))

    # จัดการการบันทึกข้อมูลฟอร์ม
    if request.method == 'POST':
        if 'save_form' in request.POST:
            # ตรวจสอบฟอร์ม UserEvaluationForm
            form = UserEvaluationForm(request.POST, instance=evaluation)
            if form.is_valid():
                form.save()
                messages.success(request, "บันทึกข้อมูลการประเมินสำเร็จ!")
            else:
                messages.error(request, "เกิดข้อผิดพลาดในการบันทึกข้อมูล กรุณาตรวจสอบข้อมูลที่กรอก")
        
        elif 'save_formset' in request.POST:
            # ตรวจสอบฟอร์ม PersonalDiagramFormset
            formset = PersonalDiagramFormset(request.POST)
            if formset.is_valid():
                formset.save()
                messages.success(request, "บันทึกข้อมูลฟอร์มสำเร็จ!")
            else:
                messages.error(request, "เกิดข้อผิดพลาดในการบันทึกข้อมูล กรุณาตรวจสอบข้อมูลที่กรอก")

    # เตรียม context เพื่อส่งไปยัง template
    context = {
        'form': form,
        'formset': formset,
        'evaluation_id': evaluation_id,
        'profile': profile,
        'selected_group': selected_group,
    }

    return render(request, 'app_evaluation/evaluation_page_from_5.html', context)

@login_required
def add_personal_diagram2(request, evaluation_id):
    # ดึง user_evaluation ตาม evaluation_id
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)

    if request.method == 'POST':
        form = PersonalDiagramForm(request.POST)
        if form.is_valid():
            personal_diagram = form.save(commit=False)
            personal_diagram.uevr_id = evaluation  # เชื่อมโยงกับ user_evaluation
            personal_diagram.save()
            messages.success(request, 'เพิ่มข้อมูลแผนพัฒนาสำเร็จ!')
            # กลับไปยังหน้า evaluation_page_5
            return redirect('evaluation_page_from_5', evaluation_id=evaluation_id)
        else:
            messages.error(request, 'เกิดข้อผิดพลาดในการบันทึก กรุณาตรวจสอบข้อมูล.')
    else:
        form = PersonalDiagramForm()

    context = {
        'form': form,
        'evaluation_id': evaluation_id,
    }
    return render(request, 'app_evaluation/add_personal_diagram2.html', context)


@login_required
def edit_personal_diagram2(request, pd_id):
    # ดึงข้อมูล PersonalDiagram ที่ต้องการแก้ไข
    personal_diagram = get_object_or_404(PersonalDiagram, pk=pd_id)

    if request.method == 'POST':
        form = PersonalDiagramForm(request.POST, instance=personal_diagram)
        if form.is_valid():
            form.save()
            messages.success(request, 'แก้ไขข้อมูลแผนพัฒนาสำเร็จ!')
            # กลับไปยังหน้า evaluation_page_5
            return redirect('evaluation_page_from_5', evaluation_id=personal_diagram.uevr_id.uevr_id)
        else:
            messages.error(request, 'เกิดข้อผิดพลาดในการแก้ไข กรุณาตรวจสอบข้อมูล.')
    else:
        form = PersonalDiagramForm(instance=personal_diagram)

    context = {
        'form': form,
        'evaluation_id': personal_diagram.uevr_id.uevr_id,
    }
    return render(request, 'app_evaluation/edit_personal_diagram2.html', context)

@login_required
def delete_personal_diagram2(request, pd_id):
    # ดึงข้อมูล PersonalDiagram ที่ต้องการลบ
    personal_diagram = get_object_or_404(PersonalDiagram, pk=pd_id)

    # ตรวจสอบคำขอเป็น POST ก่อนทำการลบ
    if request.method == 'POST':
        # เก็บ evaluation_id ก่อนลบ เพื่อ redirect กลับไปที่หน้า evaluation_page_5
        evaluation_id = personal_diagram.uevr_id.uevr_id
        personal_diagram.delete()
        messages.success(request, 'ลบข้อมูลแผนพัฒนาสำเร็จ!')
        return redirect('evaluation_page_from_5', evaluation_id=evaluation_id)

    # ในกรณีที่ไม่ใช่ POST ให้กลับไปที่ evaluation_page_5
    return redirect('evaluation_page_from_5', evaluation_id=personal_diagram.uevr_id.uevr_id)





@login_required
def view_evaluation(request, uevr_id):
    evaluation = get_object_or_404(user_evaluation, uevr_id=uevr_id)
    
    # ดึงข้อมูลที่ต้องการแสดงในหน้า "ดู" การประเมิน
    context = {
        'evaluation': evaluation,
    }

    return render(request, 'app_evaluation/evaluation_page.html', context)

@login_required
def edit_evaluation(request, evaluation_id):



    # ดึง user_evaluation ที่เลือกจาก evaluation_id
    try:
        user_evaluation_obj = user_evaluation.objects.get(uevr_id=evaluation_id)
    except user_evaluation.DoesNotExist:
        messages.error(request, "ไม่พบข้อมูลการประเมินที่เลือก")
        return redirect('search_evaluations')  # กลับไปหน้าค้นหา

    # ดึงข้อมูล user_work_info ที่เชื่อมโยงกับการประเมินนี้
    try:
        user_work = user_work_info.objects.get(user=user_evaluation_obj.user, round=user_evaluation_obj.evr_id)
    except user_work_info.DoesNotExist:
        user_work = None



    if request.method == 'POST':
        profile_form = UserProfileForm(request.POST, instance=user_evaluation_obj.user)
        extended_form = ExtendedProfileForm(request.POST, instance=user_evaluation_obj.user.profile)
        work_form = UserWorkInfoForm(request.POST, instance=user_work)

        # ตรวจสอบฟอร์มทั้งหมด
        if profile_form.is_valid() and extended_form.is_valid() and work_form.is_valid():
            # บันทึกข้อมูลผู้ใช้
            profile_form.save()
            extended_profile = extended_form.save(commit=False)
            extended_profile.user = user_evaluation_obj.user
            extended_profile.save()

            # บันทึกข้อมูลการทำงาน
            work_instance = work_form.save(commit=False)
            work_instance.user = user_evaluation_obj.user
            work_instance.round = user_evaluation_obj.evr_id
            work_instance.save()

            messages.success(request, "แก้ไขข้อมูลเรียบร้อยแล้ว!")
            return redirect('search_evaluations')  # กลับไปหน้าค้นหา
        else:
            messages.error(request, "มีข้อผิดพลาดในการบันทึกข้อมูล กรุณาตรวจสอบข้อมูลอีกครั้ง")

    else:
        # แสดงฟอร์มในกรณีที่เป็นการ GET
        profile_form = UserProfileForm(instance=user_evaluation_obj.user)
        extended_form = ExtendedProfileForm(instance=user_evaluation_obj.user.profile)
        work_form = UserWorkInfoForm(instance=user_work)

    # ตรวจสอบว่ามี uevr_id หรือไม่ เพื่อสร้าง URL สำหรับอัปโหลดไฟล์
    if user_evaluation_obj.uevr_id:
        upload_url = reverse('upload_evidence', args=[user_evaluation_obj.uevr_id])
    else:
        upload_url = None

    context = {
        'profile_form': profile_form,
        'extended_form': extended_form,
        'work_form': work_form,
        'user_evaluation': user_evaluation_obj,
        'upload_url': upload_url,  # ส่ง URL สำหรับอัปโหลดไฟล์ไปยัง template
        'evaluation_id': evaluation_id,  # ส่ง evaluation_id ผ่านไปยัง template
    }

    return render(request, 'app_evaluation/evaluation_page.html', context)




@login_required
def select_group(request):
    # ดึงข้อมูลกลุ่มที่มีอยู่ทั้งหมด
    groups = group.objects.all()
    evr_round_obj = get_evr_round()  # ดึงข้อมูลรอบการประเมินปัจจุบัน


    # ตรวจสอบว่าผู้ใช้มีการเลือกกลุ่มและมีข้อมูลการประเมินสำหรับรอบปัจจุบันแล้วหรือไม่
    selected_agreement = user_evaluation_agreement.objects.filter(
        user=request.user,
        year=timezone.now().year,
        evr_id=evr_round_obj
    ).first()

    # เพิ่มการดึงข้อมูลรอบที่ 1 (ปีที่แล้ว) ถ้าเป็นรอบที่ 2 ในปัจจุบัน
    round_1_agreement = None
    if evr_round_obj.evr_round == 2:
        round_1_agreement = user_evaluation_agreement.objects.filter(
            user=request.user,
            year=timezone.now().year - 1,
            evr_id__evr_round=1
        ).first()

    # ถ้ามีการเลือกกลุ่มแล้ว ให้ใช้ข้อมูลกลุ่มที่ถูกเลือก
    if selected_agreement:
        selected_group_name = selected_agreement.g_id.g_name
        selected_round = selected_agreement.evr_id.evr_round

        # ตรวจสอบว่ามีการสร้างการประเมินในรอบนี้แล้วหรือไม่
        evaluation = user_evaluation.objects.filter(
            user=request.user,
            evr_id=evr_round_obj
        ).first()

        if evaluation:
            # ถ้ามีข้อมูลการประเมินแล้ว ให้ไปที่หน้า evaluation_page ทันที
            return redirect('evaluation_page', evaluation_id=evaluation.uevr_id)

    # ถ้าเป็นรอบที่ 2 และมีข้อมูลกลุ่มจากรอบที่ 1 ให้ใช้กลุ่มของรอบที่ 1
    if evr_round_obj.evr_round == 2 and round_1_agreement:
        selected_group_name = round_1_agreement.g_id.g_name
        selected_round = evr_round_obj.evr_round

        # สร้างการประเมินใหม่สำหรับรอบที่ 2 โดยใช้กลุ่มจากรอบที่ 1
        evaluation = user_evaluation.objects.create(
            user=request.user,
            evr_id=evr_round_obj,
            c_gtt=None,
            c_wl=None,
            c_sumwl=None,
            approve_status=False,
            evaluater_id=None,
            evaluater_editgtt=None,
            mc_score=None,
            sc_score=None,
            adc_score=None,
            cp_num=None,
            cp_score=None,
            cp_sum=None,
            cp_main_sum=None,
            achievement_work=None,
            performing_work=None,
            other_work=None,
            sum_work=None,
            improved=None,
            suggestions=None,
        )

        # บันทึกข้อมูลการเลือกกลุ่มสำหรับรอบที่ 2
        user_evaluation_agreement.objects.update_or_create(
            user=request.user,
            year=timezone.now().year,
            evr_id=evr_round_obj,
            defaults={'g_id': round_1_agreement.g_id}
        )

        # รีไดเร็กไปที่หน้าแบบประเมินหลังจากสร้าง
        return redirect('evaluation_page', evaluation_id=evaluation.uevr_id)

    # ถ้าไม่มีการประเมิน ให้ผู้ใช้เลือกกลุ่มใหม่
    selected_group_name = None
    selected_round = evr_round_obj.evr_round

    if request.method == 'POST':
        form = GroupSelectionForm(request.POST)
        if form.is_valid():
            # เก็บค่ากลุ่มที่ผู้ใช้เลือก
            selected_group_name = form.cleaned_data['g_name']
            selected_group = get_object_or_404(group, g_name=selected_group_name)

            # สร้างหรืออัปเดตข้อมูลการเลือกกลุ่มของผู้ใช้สำหรับรอบนี้
            user_evaluation_agreement.objects.update_or_create(
                user=request.user,
                year=timezone.now().year,
                evr_id=evr_round_obj,
                defaults={'g_id': selected_group}
            )

            # สร้างแบบประเมินใหม่หลังจากเลือกกลุ่มแล้ว
            evaluation = user_evaluation.objects.create(
                user=request.user,
                evr_id=evr_round_obj,
                c_gtt=None,
                c_wl=None,
                c_sumwl=None,
                approve_status=False,
                evaluater_id=None,
                evaluater_editgtt=None,
                mc_score=None,
                sc_score=None,
                adc_score=None,
                cp_num=None,
                cp_score=None,
                cp_sum=None,
                cp_main_sum=None,
                achievement_work=None,
                performing_work=None,
                other_work=None,
                sum_work=None,
                improved=None,
                suggestions=None,
            )

            # รีไดเร็กไปที่หน้าแบบประเมินหลังจากสร้าง
            return redirect('evaluation_page', evaluation_id=evaluation.uevr_id)
    else:
        form = GroupSelectionForm()

    return render(request, 'app_evaluation/select_group.html', {
        'form': form,
        'groups': groups,
        'selected_group_name': selected_group_name,
        'selected_round': selected_round
    })



@login_required
def create_evaluation(request):
    evr_round_obj = get_evr_round()  # ใช้ฟังก์ชันเพื่อดึงรอบการประเมินที่ถูกต้อง
    if request.method == 'POST':
        form = UserEvaluationForm(request.POST)
        if form.is_valid():
            evaluation = form.save(commit=False)
            evaluation.user = request.user
            evaluation.evr_id = evr_round_obj  # เชื่อมกับรอบการประเมิน
            evaluation.save()
            return redirect('evaluation_success')
    else:
        form = UserEvaluationForm()

    return render(request, 'app_evaluation/create_evaluation.html', {'form': form, 'evr_round': evr_round_obj})



# ฟังก์ชันหน้าแบบประเมิน
@login_required
def evaluation_page(request, evaluation_id):
    user = request.user

    # ตรวจสอบว่า Profile ของผู้ใช้มีอยู่แล้วหรือไม่ ถ้าไม่ให้ redirect ไปหน้าโปรไฟล์
    try:
        profile = Profile.objects.get(user=user)
    except Profile.DoesNotExist:
        messages.error(request, "กรุณากรอกข้อมูลโปรไฟล์ก่อนทำการประเมิน")
        return redirect('profile')

    evr_round_obj = get_evr_round()  # ดึงข้อมูลรอบการประเมินปัจจุบัน

    # ดึงข้อมูล user_evaluation โดยไม่สนใจ request.user
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)

    # ดึงข้อมูล profile ที่เชื่อมกับ user_evaluation
    profile = evaluation.user.profile

    # ดึงข้อมูลการประเมินผลของผู้ใช้
    user_evaluation_agreement_obj = user_evaluation_agreement.objects.filter(user=user, evr_id=evr_round_obj).first()
    selected_group = user_evaluation_agreement_obj.g_id if user_evaluation_agreement_obj else None


    # ตรวจสอบว่า user_evaluation มีข้อมูลในรอบนี้หรือไม่
    user_evaluation_obj, created = user_evaluation.objects.get_or_create(
        user=user,
        evr_id=evr_round_obj,
        defaults={
            'c_gtt': None,
            'c_wl': None,
            'c_sumwl': None,
            'approve_status': False,
            'evaluater_id': None,
            'evaluater_editgtt': None,
            'mc_score': None,
            'sc_score': None,
            'adc_score': None,
            'cp_num': None,
            'cp_score': None,
            'cp_sum': None,
            'cp_main_sum': None,
            'achievement_work': None,
            'performing_work': None,
            'other_work': None,
            'sum_work': None,
            'improved': None,
            'suggestions': None,
            'ac_id': profile.ac_id,
            'administrative_position': profile.administrative_position,
        }
    )

    # ตรวจสอบว่าต้องอัปเดตฟิลด์ ac_id และ administrative_position หรือไม่
    if not created:  # ถ้าไม่ได้สร้างใหม่ อาจจะต้องทำการอัปเดตข้อมูล
        if user_evaluation_obj.ac_id != profile.ac_id or user_evaluation_obj.administrative_position != profile.administrative_position:
            user_evaluation_obj.ac_id = profile.ac_id
            user_evaluation_obj.administrative_position = profile.administrative_position
            user_evaluation_obj.save()

    # ดึงข้อมูลการทำงานของรอบปัจจุบัน
    user_work_current = user_work_info.objects.filter(user=user, round=evr_round_obj).first()
    # ดึงข้อมูลการทำงานในรอบที่ 1
    round_1 = evr_round.objects.filter(evr_round=1, evr_year=(timezone.now().year - 1 if evr_round_obj.evr_round == 2 else timezone.now().year)).first()

    user_work_round_1 = user_work_info.objects.filter(user=user, round=round_1).first() if round_1 else None
    
    # ดึงข้อมูลการลาในรอบปัจจุบันและรอบที่ 1
    def get_or_create_leave(user, round_obj, leave_type):
        leave, created = WorkLeave.objects.get_or_create(
            user=user,
            round=round_obj,
            leave_type=leave_type,
            defaults={'times': 0, 'days': 0}
        )
        return leave

    # สร้างหรือดึงข้อมูลการลาในแต่ละประเภทสำหรับรอบที่ 1 และรอบปัจจุบัน
    sick_leave_round_1 = get_or_create_leave(user, round_1, 'SL') if round_1 else None
    sick_leave_current = get_or_create_leave(user, evr_round_obj, 'SL')

    personal_leave_round_1 = get_or_create_leave(user, round_1, 'PL') if round_1 else None
    personal_leave_current = get_or_create_leave(user, evr_round_obj, 'PL')

    late_round_1 = get_or_create_leave(user, round_1, 'LT') if round_1 else None
    late_current = get_or_create_leave(user, evr_round_obj, 'LT')

    maternity_leave_round_1 = get_or_create_leave(user, round_1, 'ML') if round_1 else None
    maternity_leave_current = get_or_create_leave(user, evr_round_obj, 'ML')

    ordination_leave_round_1 = get_or_create_leave(user, round_1, 'OL') if round_1 else None
    ordination_leave_current = get_or_create_leave(user, evr_round_obj, 'OL')

    longsick_leave_round_1 = get_or_create_leave(user, round_1, 'LSL') if round_1 else None
    longsick_leave_current = get_or_create_leave(user, evr_round_obj, 'LSL')

    adsent_work_round_1 = get_or_create_leave(user, round_1, 'AW') if round_1 else None
    adsent_work_current = get_or_create_leave(user, evr_round_obj, 'AW')

    if request.method == 'POST':
        profile_form = UserProfileForm(request.POST, instance=user)
        extended_form = ExtendedProfileForm(request.POST, instance=profile)
        work_form_current = UserWorkInfoForm(request.POST, instance=user_work_current)

        # ฟอร์มข้อมูลการลา สำหรับรอบที่ 1 (ถ้ามี)
        if round_1:
            sick_leave_round_1_form = WorkLeaveForm(request.POST, instance=sick_leave_round_1, prefix='sick_leave_round_1')
            personal_leave_round_1_form = WorkLeaveForm(request.POST, instance=personal_leave_round_1, prefix='personal_leave_round_1')
            late_round_1_form = WorkLeaveForm(request.POST, instance=late_round_1, prefix='late_round_1')
            maternity_leave_round_1_form = WorkLeaveForm(request.POST, instance=maternity_leave_round_1, prefix='maternity_leave_round_1')
            ordination_leave_round_1_form = WorkLeaveForm(request.POST, instance=ordination_leave_round_1, prefix='ordination_leave_round_1')
            longsick_leave_round_1_form = WorkLeaveForm(request.POST, instance=longsick_leave_round_1, prefix='longsick_leave_round_1')
            adsent_work_round_1_form = WorkLeaveForm(request.POST, instance=adsent_work_round_1, prefix='adsent_work_round_1')

        # ฟอร์มข้อมูลการลา สำหรับรอบที่ 2 (ปัจจุบัน)
        sick_leave_form = WorkLeaveForm(request.POST, instance=sick_leave_current, prefix='sick_leave')
        personal_leave_form = WorkLeaveForm(request.POST, instance=personal_leave_current, prefix='personal_leave')
        late_form = WorkLeaveForm(request.POST, instance=late_current, prefix='late')
        maternity_leave_form = WorkLeaveForm(request.POST, instance=maternity_leave_current, prefix='maternity_leave')
        ordination_leave_form = WorkLeaveForm(request.POST, instance=ordination_leave_current, prefix='ordination_leave')
        longsick_leave_form = WorkLeaveForm(request.POST, instance=longsick_leave_current, prefix='longsick_leave')
        adsent_work_form = WorkLeaveForm(request.POST, instance=adsent_work_current, prefix='adsent_work')

        # ตรวจสอบฟอร์มและบันทึกเฉพาะฟอร์มที่ถูกต้อง
        if profile_form.is_valid():
            profile_form.save()
            extended_profile = extended_form.save(commit=False)
            extended_profile.user = user
            extended_profile.save()

            if work_form_current.is_valid():
                work_instance = work_form_current.save(commit=False)
                work_instance.user = user
                work_instance.round = evr_round_obj
                work_instance.save()

            # ตรวจสอบและบันทึกข้อมูลการลาในรอบที่ 1 (ถ้ามี)
            if round_1:
                form_list_round_1 = [
                    (sick_leave_round_1_form, 'SL'),
                    (personal_leave_round_1_form, 'PL'),
                    (late_round_1_form, 'LT'),
                    (maternity_leave_round_1_form, 'ML'),
                    (ordination_leave_round_1_form, 'OL'),
                    (longsick_leave_round_1_form, 'LSL'),
                    (adsent_work_round_1_form, 'AW')
                ]

                # บันทึกฟอร์มข้อมูลการลาในรอบที่ 1
                for form, leave_type in form_list_round_1:
                    if form.is_valid():
                        # กำหนดค่า leave_type
                        form.instance.leave_type = leave_type
                        form.save()
                    else:
                        print(f"{form.prefix} Form Errors: ", form.errors)

            # ตรวจสอบและบันทึกข้อมูลการลาในรอบที่ 2 (ปัจจุบัน)
            form_list_round_2 = [
                (sick_leave_form, 'SL'),
                (personal_leave_form, 'PL'),
                (late_form, 'LT'),
                (maternity_leave_form, 'ML'),
                (ordination_leave_form, 'OL'),
                (longsick_leave_form, 'LSL'),
                (adsent_work_form, 'AW')
            ]

            # บันทึกฟอร์มข้อมูลการลาในรอบที่ 2 (ปัจจุบัน)
            for form, leave_type in form_list_round_2:
                if form.is_valid():
                    # กำหนดค่า leave_type
                    form.instance.leave_type = leave_type
                    form.save()
                else:
                    print(f"{form.prefix} Form Errors: ", form.errors)

            messages.success(request, "บันทึกข้อมูลเรียบร้อยแล้ว!")
            return redirect('evaluation_page', evaluation_id=evaluation_id)
        else:
            print("Profile Form Errors: ", profile_form.errors)
            print("Extended Form Errors:", extended_form.errors)
            print("Work Form Errors:", work_form_current.errors)
            if round_1:
                for form, _ in form_list_round_1:
                    print(f"{form.prefix} Form Errors: ", form.errors)
            for form, _ in form_list_round_2:
                print(f"{form.prefix} Form Errors: ", form.errors)
            messages.error(request, "มีข้อผิดพลาดในการบันทึกข้อมูล กรุณาตรวจสอบข้อมูลอีกครั้ง")
    else:
        # ถ้าเป็นการ GET ให้แสดงฟอร์มปัจจุบัน
        profile_form = UserProfileForm(instance=user)
        extended_form = ExtendedProfileForm(instance=profile)
        work_form_current = UserWorkInfoForm(instance=user_work_current)

        # สร้างฟอร์มสำหรับ round 1 เฉพาะเมื่อ round_1 มีข้อมูล
        if round_1:
            sick_leave_round_1_form = WorkLeaveForm(instance=sick_leave_round_1, prefix='sick_leave_round_1')
            personal_leave_round_1_form = WorkLeaveForm(instance=personal_leave_round_1, prefix='personal_leave_round_1')
            late_round_1_form = WorkLeaveForm(instance=late_round_1, prefix='late_round_1')
            maternity_leave_round_1_form = WorkLeaveForm(instance=maternity_leave_round_1, prefix='maternity_leave_round_1')
            ordination_leave_round_1_form = WorkLeaveForm(instance=ordination_leave_round_1, prefix='ordination_leave_round_1')
            longsick_leave_round_1_form = WorkLeaveForm(instance=longsick_leave_round_1, prefix='longsick_leave_round_1')
            adsent_work_round_1_form = WorkLeaveForm(instance=adsent_work_round_1, prefix='adsent_work_round_1')

        sick_leave_form = WorkLeaveForm(instance=sick_leave_current, prefix='sick_leave')
        personal_leave_form = WorkLeaveForm(instance=personal_leave_current, prefix='personal_leave')
        late_form = WorkLeaveForm(instance=late_current, prefix='late')
        maternity_leave_form = WorkLeaveForm(instance=maternity_leave_current, prefix='maternity_leave')
        ordination_leave_form = WorkLeaveForm(instance=ordination_leave_current, prefix='ordination_leave')
        longsick_leave_form = WorkLeaveForm(instance=longsick_leave_current, prefix='longsick_leave')
        adsent_work_form = WorkLeaveForm(instance=adsent_work_current, prefix='adsent_work')

    # ตรวจสอบว่า uevr_id มีค่าหรือไม่เพื่อสร้าง URL สำหรับอัปโหลดไฟล์
    upload_url = reverse('upload_evidence', args=[user_evaluation_obj.uevr_id]) if user_evaluation_obj.uevr_id else None

    context = {
        'profile_form': profile_form,
        'extended_form': extended_form,
        'work_form_current': work_form_current,
        'user_work_round_1': user_work_round_1,
        'evr_round': evr_round_obj,
        'profile': profile,
        'user_evaluation_agreement': user_evaluation_agreement_obj,
        'user_evaluation_agreement_year_thai': user_evaluation_agreement_obj.year + 543,
        'selected_group': selected_group,
        'user_evaluation': user_evaluation_obj,
        'upload_url': upload_url,
        'evaluation_id': evaluation_id,
        'evaluation': evaluation,
        # ข้อมูลการลา round 1
        **({'sick_leave_round_1_form': sick_leave_round_1_form,
            'personal_leave_round_1_form': personal_leave_round_1_form,
            'late_round_1_form': late_round_1_form,
            'maternity_leave_round_1_form': maternity_leave_round_1_form,
            'ordination_leave_round_1_form': ordination_leave_round_1_form,
            'longsick_leave_round_1_form': longsick_leave_round_1_form,
            'adsent_work_round_1_form': adsent_work_round_1_form} if round_1 else {}),
        # ข้อมูลการลา round 2 (ปัจจุบัน)
        'sick_leave_form': sick_leave_form,
        'personal_leave_form': personal_leave_form,
        'late_form': late_form,
        'maternity_leave_form': maternity_leave_form,
        'ordination_leave_form': ordination_leave_form,
        'longsick_leave_form': longsick_leave_form,
        'adsent_work_form': adsent_work_form,
    }

    return render(request, 'app_evaluation/evaluation_page.html', context)


@login_required
def evaluation_page_2(request, evaluation_id):
    user = request.user
    evr_round_obj = get_evr_round()

    # ดึงข้อมูล user_evaluation ที่สัมพันธ์กับ evaluation_id
    try:
        user_evaluation_obj = user_evaluation.objects.get(pk=evaluation_id, user=user)
    except user_evaluation.DoesNotExist:
        messages.error(request, "ไม่พบข้อมูลการประเมิน")
        return redirect('select_group')

    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)

    # ดึง selected_group ที่สัมพันธ์กับ user ผ่าน user_evaluation_agreement
    try:
        selected_group = user_evaluation_agreement.objects.get(user=user, evr_id=user_evaluation_obj.evr_id).g_id
    except user_evaluation_agreement.DoesNotExist:
        selected_group = None

    # ดึงข้อมูล f_id ที่เชื่อมโยงกับ group_detail ของ selected_group
    if selected_group:
        fields = wl_field.objects.filter(group_detail__g_id=selected_group).distinct()
    else:
        fields = wl_field.objects.none()

    # ดึงข้อมูล subfield ที่ผู้ใช้เลือก
    selected_subfields = UserSelectedSubField.objects.filter(evaluation=evaluation)

    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    workload_selections = UserWorkloadSelection.objects.filter(evaluation=evaluation)
    subfields = wl_subfield.objects.filter(f_id__in=fields)
    min_workload = group.objects.filter(g_id=selected_group.g_id).values_list('g_max_workload', flat=True).first()


    # ตรวจสอบว่ามีค่า min_workload หรือไม่ ถ้าไม่มีให้ตั้งเป็นค่า default 35
    if not min_workload:
        min_workload = 35
    # คำนวณผลรวมของ calculated_workload
    # ตรวจสอบว่า workload_selections มีข้อมูลหรือไม่
    if workload_selections:
        total_workload = sum(selection.calculated_workload for selection in workload_selections)
    else:
        total_workload = 0

    # ดึง c_wl ที่สูงที่สุดจาก user_evaluation ที่สัมพันธ์กับ evr_round_obj (รอบการประเมิน)
    max_workload = user_evaluation.objects.filter( evr_id=user_evaluation_obj.evr_id).aggregate(max_workload=Max('c_wl'))['max_workload']

    # ตรวจสอบว่า max_workload เป็น None หรือไม่ ถ้าใช่ให้ตั้งค่าเป็น 0
    if max_workload is None:
        max_workload = 0

    # ตรวจสอบว่ามีคะแนนภาระงานหรือไม่
    if max_workload == 0 or total_workload <= min_workload:
        messages.error(request, f'คะแนนภาระงานต้องไม่น้อยกว่า {min_workload} ภาระงาน')

    

    # ส่วนต่างคะแนนภาระงาน (D)
    workload_difference = max_workload - total_workload

    # ตรวจสอบว่า workload_difference เป็น None หรือไม่
    if workload_difference is None:
        workload_difference = 0


    # ส่วนต่างคะแนนผลสัมฤทธิ์ของงาน (E)
    workload_difference_score = workload_difference * (28 / 115)

    # คะแนนผลสัมฤทธิ์ของงาน (F)
    achievement_work = 70 - workload_difference_score
    print(f"min_workload: {min_workload}")
    print(f"max_workload: {max_workload}")
    print(f"total_workload: {total_workload}")
    print(f"workload_difference: {workload_difference}")
    print(f"workload_difference_score: {workload_difference_score}")
    print(f"achievement_work: {achievement_work}")


    evaluation.c_wl = total_workload
    evaluation.achievement_work = round(achievement_work, 2)
    evaluation.save()

    # สร้าง formset สำหรับแต่ละ field (f_id)
    WorkloadFormset = modelformset_factory(WorkloadCriteria, fields=('c_name', 'c_num', 'c_unit', 'c_workload'), extra=0)

    formsets = {}
    for field in fields:
        formset = WorkloadFormset(queryset=WorkloadCriteria.objects.filter(f_id=field))
        formsets[field.f_id] = formset  # ใช้ f_id เป็น key ใน dictionary เพื่อเก็บ formsets แต่ละ field

    if request.method == 'POST':
        # วนลูปผ่านแต่ละ formset เพื่อตรวจสอบและบันทึกข้อมูล
        for field_id, formset in formsets.items():
            filled_formset = WorkloadFormset(request.POST, queryset=WorkloadCriteria.objects.filter(f_id__f_id=field_id))
            if filled_formset.is_valid():
                filled_formset.save()

        messages.success(request, 'บันทึกข้อมูลเรียบร้อยแล้ว!')
        return redirect('evaluation_page_2', evaluation_id=evaluation_id)

    context = {
        'user_evaluation': user_evaluation_obj,
        'formsets': formsets,  # ส่ง formsets ไปยัง template
        'fields': fields,
        'selected_subfields': selected_subfields,
        'evaluation': evaluation,
        'workload_selections': workload_selections,
        'subfields': subfields,  # ตรวจสอบให้แน่ใจว่าส่งค่า sf_id
        'total_workload': total_workload,  # ส่งผลรวมไปยังเทมเพลตเพื่อแสดง
        'achievement_work': evaluation.achievement_work,
        'evaluation_id': evaluation_id,
        'min_workload': min_workload,
    }

    return render(request, 'app_evaluation/evaluation_page_2.html', context)

@login_required
def select_subfields(request, f_id, evaluation_id):
    field = get_object_or_404(wl_field, pk=f_id)
    subfields = wl_subfield.objects.filter(f_id=field)

    # ดึงข้อมูลการประเมินที่สอดคล้องกับ evaluation_id
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)

    if request.method == 'POST':
        selected_sf_id = request.POST.get('sf_id')  # ใช้ get() เพื่อดึงค่า sf_id จากฟอร์ม
        if selected_sf_id:
            selected_sf = wl_subfield.objects.get(sf_id=selected_sf_id)
            # เพิ่มข้อมูล evaluation เข้าไปในการสร้าง UserSelectedSubField
            UserSelectedSubField.objects.create(
                user=request.user,
                evaluation=evaluation,  # เชื่อม evaluation
                f_id=field,
                sf_id=selected_sf
            )
            return redirect('evaluation_page_2', evaluation_id=evaluation_id)

    context = {
        'field': field,
        'subfields': subfields,
        'evaluation_id': evaluation_id,
    }
    return render(request, 'app_evaluation/select_subfields.html', context)

@login_required
def delete_selected_subfield(request, sf_id):
    # ตรวจสอบว่ามีการส่ง evaluation_id มาหรือไม่
    evaluation_id = request.POST.get('evaluation_id') or request.GET.get('evaluation_id')
    print(f"evaluation_id from POST: {request.POST.get('evaluation_id')}")
    print(f"evaluation_id from GET: {request.GET.get('evaluation_id')}")

    if not evaluation_id:
        messages.error(request, "ไม่พบ evaluation_id")
        return redirect('select_group')

    try:
        # ลบข้อมูลทั้งหมดใน UserSelectedSubField ที่มี sf_id
        selected_subfields = UserSelectedSubField.objects.filter(sf_id=sf_id)

        if not selected_subfields.exists():
            messages.error(request, 'ไม่พบ Subfield ที่คุณพยายามจะลบ.')
            return redirect('evaluation_page_2', evaluation_id=evaluation_id)

        if request.method == 'POST':
            # ลบข้อมูลที่ค้นหาได้ทั้งหมด
            selected_subfields.delete()
            messages.success(request, 'ลบ Subfield เรียบร้อยแล้ว!')
            return redirect('evaluation_page_2', evaluation_id=evaluation_id)

    except Exception as e:
        messages.error(request, f'เกิดข้อผิดพลาด: {str(e)}')
        return redirect('evaluation_page_2', evaluation_id=evaluation_id)

    return redirect('evaluation_page_2', evaluation_id=evaluation_id)

@login_required
def select_workload_criteria(request, evaluation_id, sf_id):
    # ดึงข้อมูล evaluation และ subfield ที่เกี่ยวข้อง
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    subfield = get_object_or_404(wl_subfield, pk=sf_id)
    
    # สร้างฟอร์มและส่งค่า subfield ไปยังฟอร์มเพื่อกรองข้อมูลใน selected_id
    criteria_form = UserWorkloadSelectionForm(initial={'sf_id': subfield}, subfield=subfield)

    # ตรวจสอบการส่งฟอร์ม
    if request.method == 'POST':
        criteria_form = UserWorkloadSelectionForm(request.POST, subfield=subfield)
        if criteria_form.is_valid():
            new_criteria = criteria_form.save(commit=False)
            new_criteria.sf_id = subfield
            new_criteria.evaluation = evaluation
            new_criteria.user = request.user

            try:
                # แปลงค่าทั้งหมดเป็น float ก่อนทำการคำนวณ
                selected_num = float(new_criteria.selected_num )
                selected_maxnum = float(new_criteria.selected_maxnum )
                selected_workload = float(new_criteria.selected_workload )

                # คำนวณค่า calculated_workload
                if selected_maxnum == 0:
                    calculated_workload = selected_num  * selected_workload
                else:
                    if selected_num <= selected_maxnum:
                        calculated_workload = selected_num * selected_workload
                    else:
                        calculated_workload = selected_maxnum  * selected_workload

                # แสดงข้อมูลการคำนวณใน log
                print(f"selected_num: {selected_num}, selected_maxnum: {selected_maxnum}, selected_workload: {selected_workload}")
                print(f"calculated_workload: {calculated_workload}")

                # บันทึกค่า calculated_workload ลงใน new_criteria
                new_criteria.calculated_workload = calculated_workload
                new_criteria.save()

                messages.success(request, f'เพิ่มภาระงานใหม่เรียบร้อยแล้ว! คำนวณภาระงาน: {calculated_workload:.2f}')
                return redirect('evaluation_page_2', evaluation_id=evaluation_id)

            except ValueError as e:
                # ถ้ามีข้อผิดพลาดในการแปลงค่า แสดงใน log
                print(f"ValueError: {e}")
                messages.error(request, 'ไม่สามารถคำนวณค่าได้ โปรดตรวจสอบข้อมูลที่กรอก.')

    # ส่งข้อมูลไปยังเทมเพลต
    context = {
        'evaluation': evaluation,
        'subfield': subfield,
        'criteria_form': criteria_form,
    }
    return render(request, 'app_evaluation/select_workload_criteria.html', context)

@login_required
def edit_workload_selection(request, selection_id):
    selection = get_object_or_404(UserWorkloadSelection, pk=selection_id)

    if request.method == 'POST':
        form = UserWorkloadSelectionForm1(request.POST, instance=selection)

        if form.is_valid():
            # ถ้าฟอร์มถูกต้อง ให้บันทึกการเปลี่ยนแปลง
            if selection.selected_maxnum == 0:
                selection.calculated_workload = selection.selected_num * selection.selected_workload
            else:
                if selection.selected_num <= selection.selected_maxnum:
                    selection.calculated_workload = selection.selected_num  * selection.selected_workload
                else:
                    selection.calculated_workload = selection.selected_maxnum  * selection.selected_workload
            form.save()
            messages.success(request, 'แก้ไขภาระงานเรียบร้อยแล้ว!')
            return redirect('evaluation_page_2', evaluation_id=selection.evaluation.uevr_id)
        else:
            messages.error(request, 'กรุณาเลือกข้อมูลภาระงานให้ครบถ้วน.')
    else:
        form = UserWorkloadSelectionForm1(instance=selection)

    return render(request, 'app_evaluation/edit_workload_selection.html', {
        'form': form,
        'selection': selection,
    })

@login_required
def delete_workload_selection(request, selection_id):
    selection = get_object_or_404(UserWorkloadSelection, pk=selection_id)

    if request.method == 'POST':
        # ลบเฉพาะการเชื่อมโยงกับ UserWorkloadSelection
        selection.delete()  
        messages.success(request, 'ลบภาระงานเรียบร้อยแล้ว!')
        return redirect('evaluation_page_2', evaluation_id=selection.evaluation.uevr_id)

    return render(request, 'app_evaluation/confirm_delete.html', {
        'selection': selection,
    })



def get_subfields(request, f_id):
    subfields = wl_subfield.objects.filter(f_id=f_id)
    subfield_list = [{'id': sf.id, 'name': sf.sf_name} for sf in subfields]
    return JsonResponse(subfield_list, safe=False)


def get_criteria(request, sf_id):
    criteria = WorkloadCriteria.objects.filter(sf_id=sf_id)
    criteria_list = [{'id': c.id, 'name': c.c_name} for c in criteria]
    return JsonResponse(criteria_list, safe=False)


@login_required
def evaluation_page_3(request, evaluation_id):
    user = request.user
    
    try:
        profile = Profile.objects.get(user=user)
    except Profile.DoesNotExist:
        messages.error(request, "กรุณากรอกข้อมูลโปรไฟล์ก่อนทำการประเมิน")
        return redirect('profile')

    evr_round_obj = get_evr_round()

    # ดึงข้อมูล user_evaluation ที่สัมพันธ์กับ evaluation_id
    try:
        user_evaluation_obj = user_evaluation.objects.get(uevr_id=evaluation_id)
    except user_evaluation.DoesNotExist:
        messages.error(request, "ไม่พบข้อมูลการประเมินที่เลือก")
        return redirect('evaluation_page', evaluation_id=evaluation_id)

    # ดึงข้อมูล competencies
    main_competencies = main_competency.objects.filter(mc_type=user_evaluation_obj.ac_id.ac_name)
    specific_competencies = specific_competency.objects.filter(sc_type=user_evaluation_obj.ac_id.ac_name)
    administrative_competencies = None  # ตั้งค่าเริ่มต้นเป็น None
    
    # เช็คว่าถ้ามีตำแหน่งบริหารและไม่ใช่ "-"
    if user_evaluation_obj.administrative_position and user_evaluation_obj.administrative_position != "-":
        administrative_competencies = administrative_competency.objects.filter()

    # ตรวจสอบการส่งข้อมูล POST
    if request.method == 'POST':
        # วนลูปผ่าน main_competencies
        for competency in main_competencies:
            actual_score = request.POST.get(f'main_actual_score_{competency.mc_id}')
            if actual_score and actual_score.isdigit():
                actual_score = int(actual_score)
                user_competency_main.objects.update_or_create(
                    evaluation=user_evaluation_obj,
                    mc_id=competency,
                    evr_id=evr_round_obj,
                    defaults={'user': user, 'umc_name': competency.mc_name, 'umc_type': competency.mc_type, 'actual_score': actual_score}
                )

        # วนลูปผ่าน specific_competencies
        for competency in specific_competencies:
            actual_score = request.POST.get(f'specific_actual_score_{competency.sc_id}')
            if actual_score and actual_score.isdigit():
                actual_score = int(actual_score)
                user_competency_councilde.objects.update_or_create(
                    evaluation=user_evaluation_obj,
                    sc_id=competency,
                    evr_id=evr_round_obj,
                    defaults={'user': user, 'ucc_name': competency.sc_name, 'ucc_type': competency.sc_type, 'actual_score': actual_score}
                )

        # วนลูปผ่าน administrative_competencies ถ้าไม่เป็น None
        if administrative_competencies is not None:
            for competency in administrative_competencies:
                actual_score = request.POST.get(f'admin_actual_score_{competency.adc_id}')
                uceo_num = request.POST.get(f'admin_uceo_num_{competency.adc_id}') 
                if actual_score and actual_score.isdigit():
                    actual_score = int(actual_score)
                    uceo_num = int(uceo_num) if uceo_num.isdigit() else 0
                    user_competency_ceo.objects.update_or_create(
                        evaluation=user_evaluation_obj,
                        adc_id=competency,
                        evr_id=evr_round_obj,
                        defaults={'user': user, 'uceo_name': competency.adc_name, 'uceo_type': competency.adc_type, 'actual_score': actual_score,'uceo_num': uceo_num }
                    )

        messages.success(request, "บันทึกคะแนนเรียบร้อยแล้ว!")
        return redirect('evaluation_page_3', evaluation_id=evaluation_id)

    # ดึงข้อมูลคะแนนที่เคยกรอก
    main_scores = user_competency_main.objects.filter(evaluation=user_evaluation_obj)
    specific_scores = user_competency_councilde.objects.filter(evaluation=user_evaluation_obj)
    administrative_scores = user_competency_ceo.objects.filter(evaluation=user_evaluation_obj)

    # เก็บจำนวนสมรรถนะที่ได้คะแนนตามเงื่อนไขต่าง ๆ
    score_count = {3: 0, 2: 0, 1: 0, 0: 0}

    # คำนวณคะแนนสำหรับ main competencies
    main_competency_total = 0
    for score in main_scores:
        calculated_score = calculate_competency_score(score.actual_score, score.mc_id.mc_num)
        score_count[calculated_score] += 1
        main_competency_total += calculated_score

    # คำนวณคะแนนสำหรับ specific competencies
    specific_competency_total = 0
    for score in specific_scores:
        calculated_score = calculate_competency_score(score.actual_score, score.sc_id.sc_num)
        score_count[calculated_score] += 1
        specific_competency_total += calculated_score

    # คำนวณคะแนนสำหรับ administrative competencies ถ้าไม่เป็น None
    administrative_competency_total = 0
    if administrative_competencies is not None:
        for score in administrative_scores:
            calculated_score = calculate_competency_score(score.actual_score, score.uceo_num)
            score_count[calculated_score] += 1
            administrative_competency_total += calculated_score

    # ผลรวมคะแนนทั้งหมด
    total_score = main_competency_total + specific_competency_total + administrative_competency_total

    # คำนวณคะแนนในแต่ละกรณี (คูณ 3, 2, 1 และ 0)
    score_3_total = score_count[3] * 3
    score_2_total = score_count[2] * 2
    score_1_total = score_count[1] * 1
    score_0_total = score_count[0] * 0

    # คำนวณค่ารวมของคะแนนที่คาดหวังสำหรับแต่ละตาราง
    main_max_num = sum([c.mc_num for c in main_competencies])
    specific_max_num = sum([c.sc_num for c in specific_competencies])
    administrative_max_num = sum([score.uceo_num for score in administrative_scores]) if administrative_competencies is not None else 0

    # คำนวณคะแนนจากสูตร
    total_max_num = main_max_num + specific_max_num + administrative_max_num

    calculated_score = 0
    if total_max_num > 0:
        calculated_score = (total_score / total_max_num) * 30


    print(f"Total Score: {total_score}")
    print(f"Total Max Num: {total_max_num}")

    # บันทึก calculated_score ใน mc_score ของ user_evaluation
    user_evaluation_obj.mc_score = calculated_score
    user_evaluation_obj.save()
    
    context = {
        'profile': profile,
        'main_competencies': main_competencies,
        'specific_competencies': specific_competencies,
        'administrative_competencies': administrative_competencies,
        'main_scores': main_scores,
        'specific_scores': specific_scores,
        'administrative_scores': administrative_scores,
        'main_competency_total': main_competency_total,
        'specific_competency_total': specific_competency_total,
        'administrative_competency_total': administrative_competency_total,
        'total_score': total_score,
        'count_3': score_count[3],
        'count_2': score_count[2],
        'count_1': score_count[1],
        'count_0': score_count[0],
        'score_3_total': score_3_total,
        'score_2_total': score_2_total,
        'score_1_total': score_1_total,
        'score_0_total': score_0_total,
        'main_max_num': main_max_num,
        'specific_max_num': specific_max_num,
        'administrative_max_num': administrative_max_num,
        'calculated_score': calculated_score,
        'evaluation_id': evaluation_id,
    }

    return render(request, 'app_evaluation/evaluation_page_3.html', context)

@login_required
def upload_evidence(request, criteria_id):
    # ดึงข้อมูล WorkloadCriteria โดยใช้ criteria_id
    criteria = get_object_or_404(UserWorkloadSelection, pk=criteria_id)
    evaluation = criteria.evaluation  # ดึงค่า evaluation จาก UserWorkloadSelection

    if request.method == 'POST':
        form = UserEvidentForm(request.POST, request.FILES)
        files = request.FILES.getlist('file')
        pictures = request.FILES.getlist('picture')

        if form.is_valid():
            # บันทึกไฟล์ PDF และ DOCX
            for file in files:
                new_filename = f"{uuid.uuid4()}_{file.name}"
                evidence = user_evident(
                    uwls_id=criteria,  # ตั้งค่า uwls_id ด้วย instance ของ UserWorkloadSelection
                    uevr_id=evaluation,  # ตั้งค่า uevr_id ด้วย instance ของ UserEvaluation
                    file=file,
                    filename=new_filename
                )
                evidence.save()

            # ลดขนาดรูปภาพก่อนบันทึก
            for picture in pictures:
                new_filename = f"{uuid.uuid4()}_{picture.name}"
                image = Image.open(picture)

                # ลดขนาดรูปภาพ
                max_size = (800, 800)  # ขนาดที่ต้องการ
                image.thumbnail(max_size)

                # บันทึกรูปภาพที่ลดขนาด
                img_io = ContentFile(b'')
                image_format = image.format or 'JPEG'
                image.save(img_io, format=image_format)

                # บันทึกไฟล์ลงใน storage
                file_name = default_storage.save(f"uploads/{new_filename}", img_io)
                evidence = user_evident(
                    uwls_id=criteria,  # ตั้งค่า uwls_id ด้วย instance ของ UserWorkloadSelection
                    uevr_id=evaluation,  # ตั้งค่า uevr_id ด้วย instance ของ UserEvaluation
                    picture=file_name,
                    filename=new_filename
                )
                evidence.save()

            messages.success(request, "อัปโหลดไฟล์เรียบร้อยแล้ว!")
            return redirect('upload_evidence', criteria_id=criteria_id)

    else:
        form = UserEvidentForm()

    # แสดงรายการไฟล์ที่อัปโหลด
    evidences = user_evident.objects.filter(uwls_id=criteria)

    return render(request, 'app_evaluation/upload_evidence.html', {
        'form': form,
        'evidences': evidences,  # ส่งรายการไฟล์/รูปภาพไปยัง template
        'criteria': criteria  # ส่งข้อมูล criteria ไปยัง template
    })

@login_required
def delete_evidence(request, evidence_id):
    evidence = get_object_or_404(user_evident, uevd_id=evidence_id)

    # ลบไฟล์หรือรูปภาพที่เกี่ยวข้องจากระบบ
    if evidence.file:
        if os.path.isfile(evidence.file.path):
            os.remove(evidence.file.path)  # ลบไฟล์จาก storage
    if evidence.picture:
        if os.path.isfile(evidence.picture.path):
            os.remove(evidence.picture.path)  # ลบรูปภาพจาก storage

    # ลบรายการ evidence จากฐานข้อมูล
    evidence.delete()
    messages.success(request, "ลบไฟล์/รูปภาพเรียบร้อยแล้ว!")
    
    # ใช้ `redirect` ให้ไปยัง `upload_evidence` โดยใช้ `criteria_id`
    return redirect('upload_evidence', criteria_id=evidence.uwls_id.id)  # เปลี่ยน `evaluation_id` เป็น `criteria_id`

@login_required
def evaluation_page_4(request, evaluation_id):
    # Get the current user
    user = request.user

    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)

    # Fetch evaluation object based on ID
    try:
        user_evaluation_obj = user_evaluation.objects.get(pk=evaluation_id, user=user)
    except user_evaluation.DoesNotExist:
        messages.error(request, "ไม่พบข้อมูลการประเมิน")
        return redirect('select_group')
    
    # ตรวจสอบการร้องขอจากผู้ใช้
    if request.method == 'POST':
        # ดึงข้อมูลจากฟอร์มที่ถูกส่งมา
        remark_achievement = request.POST.get('remark_achievement')
        remark_mc = request.POST.get('remark_mc')
        remark_other = request.POST.get('remark_other')
        remark_total = request.POST.get('remark_total')

        # บันทึกข้อมูลหมายเหตุ (อาจเพิ่มฟิลด์ใน model `user_evaluation` สำหรับบันทึกหมายเหตุ)
        evaluation.remark_achievement = remark_achievement
        evaluation.remark_mc = remark_mc
        evaluation.remark_other = remark_other
        evaluation.remark_total = remark_total

        evaluation.save()  # บันทึกการเปลี่ยนแปลง

        messages.success(request, 'บันทึกหมายเหตุเรียบร้อยแล้ว!')
        return redirect('evaluation_page_4', evaluation_id=evaluation_id)
    

    # Any specific business logic for `evaluation_page_4`
    # Example: Calculate totals, aggregate scores, etc.
    achievement_work = user_evaluation_obj.achievement_work or 0
    mc_score = user_evaluation_obj.mc_score or 0
    total_score = achievement_work + mc_score

    print(f"Achievement Work: {achievement_work}, MC Score: {mc_score}, Total Score: {total_score}")


    if total_score >= 90:
        level = 'ดีเด่น'
    elif total_score >= 80:
        level = 'ดีมาก'
    elif total_score >= 70:
        level = 'ดี'
    elif total_score >= 60:
        level = 'พอใช้'
    else:
        level = 'ต้องปรับปรุง'

    print(f"Total Score: {total_score}, Level: {level}")

    # Context to pass to template
    context = {
        'user_evaluation': user_evaluation_obj,
        'achievement_work': user_evaluation_obj.achievement_work,  # Achievement work value
        'mc_score': user_evaluation_obj.mc_score,
        'total_score': total_score,
        'level': level,
        'remark_achievement': evaluation.remark_achievement,
        'remark_mc': evaluation.remark_mc,
        'remark_other': evaluation.remark_other,
        'remark_total': evaluation.remark_total,
        'evaluation_id': evaluation_id,
        # Add more variables as needed
    }

    return render(request, 'app_evaluation/evaluation_page_4.html', context)

@login_required
def evaluation_page_5(request, evaluation_id):
    # Fetch the evaluation object
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id, user=request.user)

    # Initialize form and formset
    PersonalDiagramFormset = modelformset_factory(PersonalDiagram, fields=('skill_evol', 'dev', 'dev_time'), extra=0)
    
    form = UserEvaluationForm(instance=evaluation)
    formset = PersonalDiagramFormset(queryset=PersonalDiagram.objects.filter(uevr_id=evaluation))

    if request.method == 'POST':
        if 'save_form' in request.POST:
            # If 'save_form' button is clicked
            form = UserEvaluationForm(request.POST, instance=evaluation)
            if form.is_valid():
                form.save()
                messages.success(request, "บันทึกข้อมูลการประเมินสำเร็จ!")
            else:
                messages.error(request, "เกิดข้อผิดพลาดในการบันทึกข้อมูล กรุณาตรวจสอบข้อมูลที่กรอก")
        
        elif 'save_formset' in request.POST:
            # If 'save_formset' button is clicked
            formset = PersonalDiagramFormset(request.POST)
            if formset.is_valid():
                formset.save()
                messages.success(request, "บันทึกข้อมูลฟอร์มสำเร็จ!")
            else:
                messages.error(request, "เกิดข้อผิดพลาดในการบันทึกข้อมูล กรุณาตรวจสอบข้อมูลที่กรอก")

    context = {
        'form': form,
        'formset': formset,
        'evaluation_id': evaluation_id,
    }
    return render(request, 'app_evaluation/evaluation_page_5.html', context)


@login_required
def add_personal_diagram(request, evaluation_id):
    # ดึง user_evaluation ตาม evaluation_id
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)

    if request.method == 'POST':
        form = PersonalDiagramForm(request.POST)
        if form.is_valid():
            personal_diagram = form.save(commit=False)
            personal_diagram.uevr_id = evaluation  # เชื่อมโยงกับ user_evaluation
            personal_diagram.save()
            messages.success(request, 'เพิ่มข้อมูลแผนพัฒนาสำเร็จ!')
            # กลับไปยังหน้า evaluation_page_5
            return redirect('evaluation_page_5', evaluation_id=evaluation_id)
        else:
            messages.error(request, 'เกิดข้อผิดพลาดในการบันทึก กรุณาตรวจสอบข้อมูล.')
    else:
        form = PersonalDiagramForm()

    context = {
        'form': form,
        'evaluation_id': evaluation_id,
    }
    return render(request, 'app_evaluation/add_personal_diagram.html', context)


@login_required
def edit_personal_diagram(request, pd_id):
    # ดึงข้อมูล PersonalDiagram ที่ต้องการแก้ไข
    personal_diagram = get_object_or_404(PersonalDiagram, pk=pd_id)

    if request.method == 'POST':
        form = PersonalDiagramForm(request.POST, instance=personal_diagram)
        if form.is_valid():
            form.save()
            messages.success(request, 'แก้ไขข้อมูลแผนพัฒนาสำเร็จ!')
            # กลับไปยังหน้า evaluation_page_5
            return redirect('evaluation_page_5', evaluation_id=personal_diagram.uevr_id.uevr_id)
        else:
            messages.error(request, 'เกิดข้อผิดพลาดในการแก้ไข กรุณาตรวจสอบข้อมูล.')
    else:
        form = PersonalDiagramForm(instance=personal_diagram)

    context = {
        'form': form,
        'evaluation_id': personal_diagram.uevr_id.uevr_id,
    }
    return render(request, 'app_evaluation/edit_personal_diagram.html', context)

@login_required
def delete_personal_diagram(request, pd_id):
    # ดึงข้อมูล PersonalDiagram ที่ต้องการลบ
    personal_diagram = get_object_or_404(PersonalDiagram, pk=pd_id)

    # ตรวจสอบคำขอเป็น POST ก่อนทำการลบ
    if request.method == 'POST':
        # เก็บ evaluation_id ก่อนลบ เพื่อ redirect กลับไปที่หน้า evaluation_page_5
        evaluation_id = personal_diagram.uevr_id.uevr_id
        personal_diagram.delete()
        messages.success(request, 'ลบข้อมูลแผนพัฒนาสำเร็จ!')
        return redirect('evaluation_page_5', evaluation_id=evaluation_id)

    # ในกรณีที่ไม่ใช่ POST ให้กลับไปที่ evaluation_page_5
    return redirect('evaluation_page_5', evaluation_id=personal_diagram.uevr_id.uevr_id)

@login_required
def evaluation_page_6(request):
    return render(request,'app_evaluation/evaluation_page_6.html')

@login_required
def update_evr_status(request, evaluation_id):
    # Fetch the user_evaluation object
    user_evaluation_obj = get_object_or_404(user_evaluation, pk=evaluation_id)

    # Fetch the related evr_round object
    evr_round_obj = user_evaluation_obj.evr_id

    # Check request method and update evr_status
    if request.method == 'POST':
        evr_round_obj.evr_status = True  # or the desired status
        evr_round_obj.save()
        messages.success(request, 'สถานะการประเมินถูกอัพเดตเรียบร้อยแล้ว')
    
    # Redirect back to the evaluation page
    return redirect('search_evaluations')

@login_required
def update_evr_status2(request, evaluation_id):
    # Fetch the user_evaluation object
    user_evaluation_obj = get_object_or_404(user_evaluation, pk=evaluation_id)

    # Fetch the related evr_round object
    evr_round_obj = user_evaluation_obj.evr_id

    # Check request method and update evr_status
    if request.method == 'POST':
        evr_round_obj.evr_status = True  # or the desired status
        evr_round_obj.save()
        messages.success(request, 'สถานะการประเมินถูกอัพเดตเรียบร้อยแล้ว')
    
    # Redirect back to the evaluation page
    return redirect('search_evaluations_2')

@login_required
def toggle_approve_status(request, evaluation_id):
    # ดึงข้อมูล user_evaluation จาก evaluation_id
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    
    # เปลี่ยนสถานะ approve_status
    evaluation.approve_status = not evaluation.approve_status
    evaluation.save()
    
    # ส่งข้อความแสดงผลสำเร็จ
    messages.success(request, f"เปลี่ยนสถานะการอนุมัติสำเร็จเป็น {'อนุมัติ' if evaluation.approve_status else 'ไม่อนุมัติ'}")
    
    # กลับไปยังหน้าที่ต้องการ (เช่น หน้ารายการการประเมิน)
    return redirect('search_evaluations_2')  # แก้ไข URL name ให้ตรงกับ project ของคุณ



@login_required
def eval_2(request):
    return render(request,'app_evaluation/evaluation2.html')

@login_required
def eval_3(request):
    return render(request,'app_evaluation/evaluation3.html')

@login_required
def eval_4(request):
    return render(request,'app_evaluation/evaluation4.html')

@login_required
def eval_5(request):
    return render(request,'app_evaluation/evaluation5.html')





@login_required
def export_html_to_excel(request, evaluation_id):
    user = request.user
    # ดึงข้อมูล evaluation
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    profile = evaluation.user.profile
    evr_round_obj = get_evr_round()  # ดึงข้อมูลรอบการประเมินปัจจุบัน

    # ดึงข้อมูลของรอบที่ 1
    round_1 = evr_round.objects.filter(
        evr_round=1,
        evr_year=(timezone.now().year - 1 if evr_round_obj.evr_round == 2 else timezone.now().year)
    ).first()

    # ดึงข้อมูลการลาในรอบปัจจุบันและรอบที่ 1
    def get_or_create_leave(user, round_obj, leave_type):
        leave, created = WorkLeave.objects.get_or_create(
            user=user,
            round=round_obj,
            leave_type=leave_type,
            defaults={'times': 0, 'days': 0}
        )
        return leave

    # สร้างหรือดึงข้อมูลการลาในแต่ละประเภทสำหรับรอบที่ 1 และรอบปัจจุบัน
    sick_leave_round_1 = get_or_create_leave(user, round_1, 'SL') if round_1 else None
    sick_leave_current = get_or_create_leave(user, evr_round_obj, 'SL')

    personal_leave_round_1 = get_or_create_leave(user, round_1, 'PL') if round_1 else None
    personal_leave_current = get_or_create_leave(user, evr_round_obj, 'PL')

    late_round_1 = get_or_create_leave(user, round_1, 'LT') if round_1 else None
    late_current = get_or_create_leave(user, evr_round_obj, 'LT')

    maternity_leave_round_1 = get_or_create_leave(user, round_1, 'ML') if round_1 else None
    maternity_leave_current = get_or_create_leave(user, evr_round_obj, 'ML')

    ordination_leave_round_1 = get_or_create_leave(user, round_1, 'OL') if round_1 else None
    ordination_leave_current = get_or_create_leave(user, evr_round_obj, 'OL')

    longsick_leave_round_1 = get_or_create_leave(user, round_1, 'LSL') if round_1 else None
    longsick_leave_current = get_or_create_leave(user, evr_round_obj, 'LSL')

    adsent_work_round_1 = get_or_create_leave(user, round_1, 'AW') if round_1 else None
    adsent_work_current = get_or_create_leave(user, evr_round_obj, 'AW')


    # สร้าง Workbook ใหม่
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ข้อมูลการประเมิน"

    # จัดรูปแบบสำหรับหัวตาราง
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # ใส่ข้อมูลส่วนบุคคลในเซลล์ต่างๆ
    ws.merge_cells('A1:B1')
    ws['A1'] = "ชื่อ - สกุล"
    ws['A1'].alignment = center_alignment
    ws['A1'].border = thin_border

    ws.merge_cells('C1:D1')
    ws['C1'] = f"{profile.user.first_name} {profile.user.last_name}"
    ws['C1'].alignment = center_alignment
    ws['C1'].border = thin_border

    ws['A2'] = "ตำแหน่งบริหาร"
    ws.merge_cells('C2:D2')
    ws['C2'] = profile.administrative_position or "-"

    # ใส่ข้อมูลเงินเดือน
    ws['A3'] = "เงินเดือน"
    ws.merge_cells('B3:C3')
    ws['B3'] = f"{profile.salary} บาท"
    ws.merge_cells('D3:E3')
    ws['D3'] = f"เลขที่ประจำตำแหน่ง {profile.position_number}"

    # ใส่ข้อมูลสังกัด
    ws['A4'] = "สังกัด"
    ws.merge_cells('C4:E4')
    ws['C4'] = profile.affiliation

    # ใส่ข้อมูลรวมเวลาราชการ
    ws['A5'] = "รวมเวลาราชการ"
    ws.merge_cells('C5:E5')
    ws['C5'] = f"{profile.years_of_service} ปี {profile.months_of_service} เดือน {profile.days_of_service} วัน"

    # ใส่หัวตารางข้อมูลการลา
    ws['A7'] = "บันทึกการลา"
    ws.merge_cells('A7:E7')
    ws['A7'].alignment = center_alignment
    ws['A7'].border = thin_border

    # สร้างตารางสำหรับข้อมูลการลา
    ws['A8'] = "ประเภท"
    ws['B8'] = "รอบที่ 1"
    ws['C8'] = ""
    ws['D8'] = "รอบที่ 2"
    ws['E8'] = ""
    ws.merge_cells('B8:C8')
    ws.merge_cells('D8:E8')
    ws['B8'].alignment = center_alignment
    ws['D8'].alignment = center_alignment

    ws['B9'] = "ครั้ง"
    ws['C9'] = "วัน"
    ws['D9'] = "ครั้ง"
    ws['E9'] = "วัน"

    # ใส่ข้อมูลการลาแต่ละประเภท
    leave_types = [
        ("ลาป่วย", sick_leave_round_1, sick_leave_current),
        ("ลากิจ", personal_leave_round_1, personal_leave_current),
        ("มาสาย", late_round_1, late_current),
        ("ลาคลอดบุตร", maternity_leave_round_1, maternity_leave_current),
        ("ลาอุปสมบท", ordination_leave_round_1, ordination_leave_current),
        ("ลาป่วยจำเป็นต้องรักษาตัวเป็นเวลานาน", longsick_leave_round_1, longsick_leave_current),
        ("ขาดราชการ", adsent_work_round_1, adsent_work_current),
    ]

    row = 10
    for leave_type, round_1, current_round in leave_types:
        ws[f'A{row}'] = leave_type
        ws[f'B{row}'] = round_1.times if round_1 else ""
        ws[f'C{row}'] = round_1.days if round_1 else ""
        ws[f'D{row}'] = current_round.times if current_round else ""
        ws[f'E{row}'] = current_round.days if current_round else ""
        row += 1

    # จัดการเส้นขอบและการจัดตำแหน่ง
    for row in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border

    # บันทึกข้อมูลลงใน BytesIO แทนที่จะเป็นไฟล์จริง
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # สร้าง response สำหรับการดาวน์โหลดไฟล์ Excel
    response = HttpResponse(
        excel_file,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="evaluation_{evaluation_id}.xlsx"'
    return response


@login_required
def export_evaluation2_to_excel(request, evaluation_id):
    user = request.user
    evr_round_obj = get_evr_round()

    # ดึงข้อมูล user_evaluation ที่สัมพันธ์กับ evaluation_id
    try:
        user_evaluation_obj = user_evaluation.objects.get(pk=evaluation_id, user=user)
    except user_evaluation.DoesNotExist:
        messages.error(request, "ไม่พบข้อมูลการประเมิน")
        return redirect('select_group')

    # ดึง selected_group ที่สัมพันธ์กับ user ผ่าน user_evaluation_agreement
    try:
        selected_group = user_evaluation_agreement.objects.get(user=user, evr_id=user_evaluation_obj.evr_id).g_id
    except user_evaluation_agreement.DoesNotExist:
        selected_group = None

    # ดึงข้อมูล f_id ที่เชื่อมโยงกับ group_detail ของ selected_group
    if selected_group:
        fields = wl_field.objects.filter(group_detail__g_id=selected_group).distinct()
    else:
        fields = wl_field.objects.none()

    # ดึงข้อมูล subfield ที่ผู้ใช้เลือก
    selected_subfields = UserSelectedSubField.objects.filter(evaluation=user_evaluation_obj)

    # ดึงข้อมูล workload_selections
    workload_selections = UserWorkloadSelection.objects.filter(evaluation=user_evaluation_obj)
    subfields = wl_subfield.objects.filter(f_id__in=fields)

    # ดึงค่า min_workload และ total_workload
    min_workload = group.objects.filter(g_id=selected_group.g_id).values_list('g_max_workload', flat=True).first() or 35
    total_workload = sum(selection.calculated_workload for selection in workload_selections) or 0

    # คำนวณ achievement_work
    max_workload = user_evaluation.objects.filter(evr_id=user_evaluation_obj.evr_id).aggregate(max_workload=Max('c_wl'))['max_workload'] or 0
    workload_difference = max_workload - total_workload
    workload_difference_score = workload_difference * (28 / 115)
    achievement_work = 70 - workload_difference_score

    # สร้างไฟล์ Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"

    # จัดรูปแบบ
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    # หัวข้อและข้อมูลในตาราง
    headers = ["Field", "Subfield", "Workload", "Achievement Work", "Total Workload", "Min Workload"]
    ws.append(headers)

    # เพิ่มข้อมูลลงในแถว
    for field in fields:
        for subfield in subfields:
            ws.append([field.f_name, subfield.sf_name, total_workload, achievement_work, total_workload, min_workload])

    # ใส่ข้อมูลอื่นๆ เช่น ผลสัมฤทธิ์ของงาน
    ws.append(["Total Workload", total_workload])
    ws.append(["Achievement Work", achievement_work])

    # จัดการเส้นขอบและจัดตำแหน่ง
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border

    # บันทึกข้อมูลลงใน BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # สร้าง response เพื่อดาวน์โหลดไฟล์ Excel
    response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="evaluation_{evaluation_id}.xlsx"'
    return response


@login_required
def export_evaluation3_to_excel(request, evaluation_id):
    user = request.user

    # ดึงข้อมูล profile
    try:
        profile = Profile.objects.get(user=user)
    except Profile.DoesNotExist:
        messages.error(request, "ไม่พบข้อมูลโปรไฟล์")
        return redirect('profile')

    # ดึงข้อมูล user_evaluation ที่สัมพันธ์กับ evaluation_id
    try:
        user_evaluation_obj = user_evaluation.objects.get(uevr_id=evaluation_id)
    except user_evaluation.DoesNotExist:
        messages.error(request, "ไม่พบข้อมูลการประเมินที่เลือก")
        return redirect('evaluation_page', evaluation_id=evaluation_id)

    # ดึงข้อมูล competencies
    main_competencies = main_competency.objects.filter(mc_type=user_evaluation_obj.ac_id.ac_name)
    specific_competencies = specific_competency.objects.filter(sc_type=user_evaluation_obj.ac_id.ac_name)
    administrative_competencies = None
    
    if user_evaluation_obj.administrative_position and user_evaluation_obj.administrative_position != "-":
        administrative_competencies = administrative_competency.objects.all()

    # ดึงข้อมูลคะแนนที่เคยกรอก
    main_scores = user_competency_main.objects.filter(evaluation=user_evaluation_obj)
    specific_scores = user_competency_councilde.objects.filter(evaluation=user_evaluation_obj)
    administrative_scores = user_competency_ceo.objects.filter(evaluation=user_evaluation_obj)

    # สร้างไฟล์ Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"

    # จัดรูปแบบ
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    # หัวตาราง
    headers = ["Competency", "Expected Level", "Actual Level", "Competency Type"]
    ws.append(headers)

    # ใส่ข้อมูล Main Competencies
    for score in main_scores:
        ws.append([score.mc_id.mc_name, score.mc_id.mc_num, score.actual_score, "Main Competency"])

    # ใส่ข้อมูล Specific Competencies
    for score in specific_scores:
        ws.append([score.sc_id.sc_name, score.sc_id.sc_num, score.actual_score, "Specific Competency"])

    # ใส่ข้อมูล Administrative Competencies (ถ้ามี)
    if administrative_competencies:
        for score in administrative_scores:
            ws.append([score.adc_id.adc_name, score.uceo_num, score.actual_score, "Administrative Competency"])

    # ใส่คะแนนรวมและคะแนนคำนวณ
    ws.append(["", "", "", ""])  # เพิ่มช่องว่างระหว่างตารางและผลรวม
    ws.append(["Total Score", user_evaluation_obj.mc_score])

    # จัดการเส้นขอบและการจัดตำแหน่ง
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border

    # บันทึกข้อมูลลงใน BytesIO แทนที่จะเป็นไฟล์จริง
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # สร้าง response เพื่อดาวน์โหลดไฟล์ Excel
    response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="evaluation_{evaluation_id}.xlsx"'
    return response


@login_required
def export_evaluation_page_4_to_excel(request, evaluation_id):
    # ดึงข้อมูลการประเมินจาก user_evaluation
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id)
    user = request.user
    
    # ดึงข้อมูลจาก Profile
    try:
        profile = Profile.objects.get(user=user)
    except Profile.DoesNotExist:
        messages.error(request, "ไม่พบข้อมูลโปรไฟล์")
        return redirect('profile')

    # คะแนนผลสัมฤทธิ์ของงานและสมรรถนะ
    achievement_work = evaluation.achievement_work or 0
    mc_score = evaluation.mc_score or 0
    total_score = achievement_work + mc_score

    # ตรวจสอบระดับผลการประเมิน
    if total_score >= 90:
        level = 'ดีเด่น'
    elif total_score >= 80:
        level = 'ดีมาก'
    elif total_score >= 70:
        level = 'ดี'
    elif total_score >= 60:
        level = 'พอใช้'
    else:
        level = 'ต้องปรับปรุง'

    # สร้างไฟล์ Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"

    # จัดรูปแบบเซลล์
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    # หัวตาราง
    headers = ["องค์ประกอบการประเมิน", "คะแนนเต็ม", "คะแนนที่ได้", "หมายเหตุ"]
    ws.append(headers)

    # องค์ประกอบที่ 1: ผลสัมฤทธิ์ของงาน
    ws.append([
        "องค์ประกอบที่ 1: ผลสัมฤทธิ์ของงาน",
        70,
        achievement_work,
        evaluation.remark_achievement or ""
    ])

    # องค์ประกอบที่ 2: สมรรถนะ
    ws.append([
        "องค์ประกอบที่ 2: พฤติกรรมการปฏิบัติราชการ (สมรรถนะ)",
        30,
        mc_score,
        evaluation.remark_mc or ""
    ])

    # องค์ประกอบอื่น ๆ
    ws.append([
        "องค์ประกอบอื่น ๆ",
        "",
        "",
        evaluation.remark_other or ""
    ])

    # รวมคะแนน
    ws.append([
        "รวม",
        100,
        total_score,
        evaluation.remark_total or ""
    ])

    # ใส่ระดับผลการประเมิน
    ws.append(["ระดับผลการประเมิน", "", "", level])

    # จัดการเส้นขอบและการจัดตำแหน่ง
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border

    # บันทึกข้อมูลลงใน BytesIO แทนที่จะเป็นไฟล์จริง
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # สร้าง response เพื่อดาวน์โหลดไฟล์ Excel
    response = HttpResponse(
        excel_file,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="evaluation_{evaluation_id}.xlsx"'
    return response


@login_required
def export_evaluation_page_5_to_excel(request, evaluation_id):
    # ดึงข้อมูลการประเมิน
    evaluation = get_object_or_404(user_evaluation, pk=evaluation_id, user=request.user)
    
    # ดึงข้อมูล PersonalDiagram ที่สัมพันธ์กับการประเมิน
    personal_diagrams = PersonalDiagram.objects.filter(uevr_id=evaluation)

    # สร้าง Workbook ใหม่
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Page 5"

    # จัดรูปแบบเซลล์
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    # เพิ่มหัวตาราง
    headers = ["ทักษะ/สมรรถนะ", "วิธีการพัฒนา", "ช่วงเวลา", "หมายเหตุ"]
    ws.append(headers)

    # ใส่ข้อมูลลงใน Excel จาก PersonalDiagram
    for diagram in personal_diagrams:
        ws.append([
            diagram.skill_evol or "",
            diagram.dev or "",
            diagram.dev_time or "",
            ""
        ])

    # เพิ่มข้อมูลส่วน "ความเห็นเพิ่มเติมของผู้ประเมิน" จากฟอร์ม UserEvaluation
    ws.append([])
    ws.append(["ความเห็นเพิ่มเติมของผู้ประเมิน"])
    ws.append(["จุดเด่น/สิ่งที่ควรปรับปรุง:", evaluation.improved or ""])
    ws.append(["ข้อเสนอแนะในการพัฒนา:", evaluation.suggestions or ""])

    # จัดการเส้นขอบและการจัดตำแหน่ง
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border

    # บันทึกข้อมูลลงใน BytesIO แทนที่จะเป็นไฟล์จริง
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # สร้าง response เพื่อดาวน์โหลดไฟล์ Excel
    response = HttpResponse(
        excel_file,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="evaluation_page_5_{evaluation_id}.xlsx"'
    return response


@login_required
def export_evaluation_page_6_to_excel(request):
    # สร้าง Workbook และ worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Page 6"

    # กำหนดรูปแบบสำหรับเซลล์
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    # เพิ่มส่วนที่ 5 การรับทราบผลการประเมิน
    ws.append(['ส่วนที่ 5 การรับทราบผลการประเมิน'])
    ws.append(['ผู้รับการประเมิน :'])
    ws.append(['ได้รับทราบผลการประเมินและแผนพัฒนา', 'ลงชื่อ..........................................................'])
    ws.append(['การปฏิบัติราชการรายบุคคลแล้ว', 'ตำแหน่ง.....................................................'])
    ws.append(['', 'วันที่...........................................................'])
    ws.append([])  # เว้นบรรทัด

    ws.append(['ผู้ประเมิน :'])
    ws.append(['ได้แจ้งผลการประเมินและผู้รับการประเมิน', 'ลงชื่อ..........................................................'])
    ws.append(['ได้ลงนามรับทราบ', 'ตำแหน่ง...................................................'])
    ws.append([])  # เว้นบรรทัด

    ws.append(['ได้แจ้งผลการประเมินเมื่อวันที่.................................................', 'วันที่..........................................................'])
    ws.append(['แต่ผู้รับการประเมินไม่ลงนามรับทราบผลการประเมิน'])
    ws.append(['โดยมี.......................................................................เป็นพยาน'])
    ws.append([])  # เว้นบรรทัด

    # เพิ่มส่วนที่ 6 ความเห็นของผู้บังคับบัญชาเหนือขึ้นไป
    ws.append(['ส่วนที่ 6 ความเห็นของผู้บังคับบัญชาเหนือขึ้นไป'])
    ws.append(['ผู้บังคับบัญชาเหนือขึ้นไป :'])
    ws.append(['เห็นด้วยกับผลการประเมิน', 'ลงชื่อ..........................................................'])
    ws.append([])  # เว้นบรรทัด
    ws.append(['มีความเห็นแตกต่าง ดังนี้'])
    ws.append(['.........................................................................................................'])
    ws.append(['.........................................................................................................', 'ตำแหน่ง......................................................'])
    ws.append(['วันที่............................................................'])

    # เพิ่มผู้บังคับบัญชาอีกชั้นหนึ่ง (ถ้ามี)
    ws.append([])  # เว้นบรรทัด
    ws.append(['ผู้บังคับบัญชาเหนือขึ้นไปอีกชั้นหนึ่ง (ถ้ามี) :'])
    ws.append(['เห็นด้วยกับผลการประเมิน', 'ลงชื่อ..........................................................'])
    ws.append([])  # เว้นบรรทัด
    ws.append(['มีความเห็นแตกต่าง ดังนี้'])
    ws.append(['.........................................................................................................'])
    ws.append(['.........................................................................................................', 'ตำแหน่ง......................................................'])
    ws.append(['วันที่............................................................'])

    # จัดการเส้นขอบและการจัดตำแหน่งให้ทุกเซลล์ใน worksheet
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border

    # บันทึกข้อมูลลงใน BytesIO แทนการบันทึกเป็นไฟล์จริง
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # สร้าง response สำหรับการดาวน์โหลดไฟล์ Excel
    response = HttpResponse(
        excel_file,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="evaluation_page_6.xlsx"'
    return response




